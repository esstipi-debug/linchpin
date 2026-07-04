"""Excel replenishment agent job: read a client's planilla -> plan -> staged write-back.

The Excel twin of ``jobs/odoo_job.py``: the client's spreadsheet IS the system of
record. ``prepare`` reads the planilla (auto-detecting the sheet, the header row and
the SKU / stock / reorder-point / demand columns, Spanish or English, accents
folded), ``run`` plans the restock and STAGES the recommended order quantities as a
dry-run changeset through the safe-staging plane (``src/connectors/excel.py`` —
drift check, atomic write, backup, rollback), and the result is presented as >=2
ranked executable options honouring the never-unprotected contract. Nothing touches
the client's file until a human approves and applies the staged changeset.

Two planning modes, picked from what the planilla actually has:
- ``demand-cover``: a demand column exists -> order up to ``cover_periods`` of demand
  (same target logic as ``src/connectors/replenish.py``);
- ``reorder-point``: only a reorder-point column exists -> classic min/max: order when
  on-hand < ROP, up to ``order_up_to_factor * ROP``.

Safety-of-plan properties (each earned by an adversarial-review finding):
- values are read ``data_only`` so formula-driven cells use their Excel-cached value;
  rows whose stock still isn't numeric are SKIPPED AND COUNTED (never silently);
- rows with no planning signal (blank ROP / blank demand) are EXCLUDED AND COUNTED —
  a blank never coalesces to 0 and orders (or suppresses) anything;
- duplicate SKUs fail closed (the staged write per SKU would be ambiguous);
- column binding is deterministic (priority-ordered candidates, first label wins);
- the changeset carries GUARD cells (stock + signal, staged as no-op before==after)
  so the connector's drift check also refuses when the plan's INPUTS changed between
  staging and the human apply — not just its output cells;
- the idempotency key derives from the staged content, so re-running next week (a
  different plan) never collides with last week's apply, while re-staging the SAME
  plan stays idempotent.
"""

from __future__ import annotations

import hashlib
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src import writeback
from src.connectors.excel import ExcelWorkbookStore
from src.deliverable import DataSource, Deliverable, Finding, Kpi
from src.export import write_summary_csv
from src.guided import ExecutionOption, GuidedOutcome, as_options, verify_guided

_EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")
_HEADER_SCAN_ROWS = 20
DEFAULT_ORDER_COLUMN = "Pedir (Linchpin)"

# Priority-ORDERED candidates (tuples, not sets): when a header row carries two
# plausible labels for the same role ("Stock" AND "Existencias"), the earlier
# candidate wins — deterministically, across processes and hash seeds.
_SKU_CANDIDATES = (
    "sku", "codigo", "codigo sku", "product id", "product_id", "producto", "product",
    "code", "item", "articulo", "material", "referencia", "ref", "skus",
)
_STOCK_CANDIDATES = (
    "stock actual", "stock", "on hand", "on-hand", "on_hand", "existencia",
    "existencias", "inventario", "disponible", "cantidad", "qty", "quantity", "saldo",
)
_ROP_CANDIDATES = (
    "punto de reorden", "punto reorden", "reorder point", "rop", "stock minimo",
    "punto de pedido", "punto pedido", "minimo", "min", "reorden",
)
_DEMAND_CANDIDATES = (
    "demanda semanal", "demanda mensual", "demanda", "weekly demand", "demand",
    "venta promedio", "ventas promedio", "average sales", "avg sales",
    "pronostico", "forecast", "consumo promedio", "consumo",
)


def _fold(label: object) -> str:
    """Lowercased, accent-stripped, whitespace-collapsed header label."""
    text = unicodedata.normalize("NFKD", str(label)).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.lower().split())


def _num(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


@dataclass(frozen=True)
class PlanillaRow:
    row: int
    sku: str
    on_hand: float
    reorder_point: float | None
    demand_per_period: float | None


@dataclass(frozen=True)
class ReplenishmentLine:
    sku: str
    on_hand: float
    target: float
    restock_qty: float


@dataclass(frozen=True)
class ExcelReplenishmentReport:
    filename: str
    sheet: str
    mode: str  # "demand-cover" | "reorder-point"
    lines: tuple[ReplenishmentLine, ...]
    restock: dict[str, float]
    n_skus: int
    n_restock: int
    n_unplanned: int          # rows excluded: no planning signal (blank ROP/demand)
    n_skipped_rows: int       # rows excluded: SKU present but stock not numeric
    total_restock: float
    cover_periods: float
    order_up_to_factor: float
    changeset: writeback.Changeset | None
    outcome: GuidedOutcome
    summary: str


def _scan_sheet(ws, wanted: dict[str, tuple[str, ...]]) -> tuple[int, dict[str, str]] | None:
    """First row (within the scan window) containing a SKU-candidate header.

    Returns (header_row, {role: column_letter}) or None. When the same label
    appears twice in a row, the LEFTMOST column wins (setdefault).
    """
    for row in ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS):
        labels: dict[str, str] = {}
        for c in row:
            if c.value is not None:
                labels.setdefault(_fold(c.value), c.column_letter)
        found: dict[str, str] = {}
        for role, candidates in wanted.items():
            for cand in candidates:
                if cand in labels:
                    found[role] = labels[cand]
                    break
        if "sku" in found:
            return row[0].row, found
    return None


def _qualifies(cols: dict[str, str]) -> bool:
    return "stock" in cols and ("rop" in cols or "demand" in cols)


def prepare(data_path: str | None, params: dict | None = None) -> dict:
    """Read the client's planilla and resolve where everything lives.

    ``params`` overrides beat auto-detection: ``sheet``, ``sku_column``,
    ``stock_column``, ``rop_column``, ``demand_column``, ``order_column``.
    """
    params = params or {}
    if not data_path:
        raise ValueError("an Excel file (.xlsx/.xlsm) is required")
    path = Path(data_path)
    if path.suffix.lower() not in _EXCEL_SUFFIXES:
        raise ValueError(
            f"excel_replenishment needs an Excel file (.xlsx/.xlsm), got {path.suffix!r} - "
            "for CSV demand data use the inventory_optimization tool instead"
        )
    try:
        # data_only: formula cells read their Excel-cached VALUE (what the client
        # sees), not the formula text - numbers stay numbers for planning.
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        raise
    except Exception as exc:  # zip/corruption/encryption errors are library-specific
        raise ValueError(
            f"could not open {path.name} as an Excel workbook ({type(exc).__name__}: {exc}) - "
            "is it a real, unencrypted .xlsx/.xlsm file?"
        ) from exc

    order_column = str(params.get("order_column", DEFAULT_ORDER_COLUMN))
    wanted: dict[str, tuple[str, ...]] = {
        "sku": (_fold(params["sku_column"]),) if params.get("sku_column") else _SKU_CANDIDATES,
        "stock": (_fold(params["stock_column"]),) if params.get("stock_column") else _STOCK_CANDIDATES,
        "rop": (_fold(params["rop_column"]),) if params.get("rop_column") else _ROP_CANDIDATES,
        "demand": (_fold(params["demand_column"]),) if params.get("demand_column") else _DEMAND_CANDIDATES,
        "order": (_fold(order_column),),
    }

    if params.get("sheet"):
        if params["sheet"] not in wb.sheetnames:
            raise ValueError(f"sheet {params['sheet']!r} not found (sheets: {', '.join(wb.sheetnames)})")
        sheet_names = [params["sheet"]]
    else:
        sheet_names = wb.sheetnames

    resolved = None
    partial = None  # first sheet with a SKU header but missing stock/signal, for the error
    for name in sheet_names:
        hit = _scan_sheet(wb[name], wanted)
        if hit is None:
            continue
        header_row, cols = hit
        if _qualifies(cols):
            resolved = (name, header_row, cols)
            break
        if partial is None:
            partial = (name, cols)
    if resolved is None:
        if partial is not None:
            name, cols = partial
            missing = ("a stock/on-hand column" if "stock" not in cols
                       else "a reorder-point (e.g. 'Punto Reorden') or demand (e.g. 'Demanda Semanal') column")
            raise ValueError(
                f"sheet {name!r} has a SKU column but no {missing} in the same header row - "
                "add it or pass stock_column/rop_column/demand_column explicitly"
            )
        raise ValueError(
            "could not find a SKU/product column in the first rows of any sheet - "
            "name it (e.g. 'Codigo'/'SKU') or pass sku_column/sheet explicitly"
        )
    sheet, header_row, cols = resolved
    mode = "demand-cover" if "demand" in cols else "reorder-point"

    ws = wb[sheet]
    rows: list[PlanillaRow] = []
    n_skipped = 0
    seen_skus: dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku_raw = ws[f"{cols['sku']}{r}"].value
        if sku_raw is None:
            continue  # blank/separator rows
        on_hand = _num(ws[f"{cols['stock']}{r}"].value)
        if on_hand is None:
            n_skipped += 1  # SKU present but stock not numeric (e.g. uncached formula)
            continue
        sku = str(sku_raw).strip()
        if sku in seen_skus:
            # Two rows claiming the same SKU make "the" write-back cell ambiguous -
            # same fail-closed stance as the connector's resolve_row_edits.
            raise ValueError(
                f"duplicate SKU {sku!r} in {sheet!r} (rows {seen_skus[sku]} and {r}) - "
                "resolve the duplicate before planning a write-back"
            )
        seen_skus[sku] = r
        rows.append(PlanillaRow(
            row=r,
            sku=sku,
            on_hand=on_hand,
            reorder_point=_num(ws[f"{cols['rop']}{r}"].value) if "rop" in cols else None,
            demand_per_period=_num(ws[f"{cols['demand']}{r}"].value) if "demand" in cols else None,
        ))
    if not rows:
        raise ValueError(
            f"no usable SKU rows under the {sheet!r} headers"
            + (f" - {n_skipped} row(s) had a non-numeric stock value; if the stock column is "
               "formula-driven, open and save the file in Excel once so values are cached"
               if n_skipped else "")
        )

    order_exists = "order" in cols
    order_letter = cols["order"] if order_exists else get_column_letter(ws.max_column + 1)
    # An optional persistent ledger makes idempotency + rollback survive across
    # processes - what the operator apply CLI needs (in-memory audit dies with it).
    ledger = None
    if params.get("ledger_path"):
        from src.writeback_store import SqliteAuditLedger
        ledger = SqliteAuditLedger(params["ledger_path"])
    return {
        "store": ExcelWorkbookStore(path, ledger=ledger),
        "filename": path.name,
        "sheet": sheet,
        "header_row": header_row,
        "mode": mode,
        "rows": tuple(rows),
        "n_skipped_rows": n_skipped,
        "input_cols": {"stock": cols["stock"], "rop": cols.get("rop"), "demand": cols.get("demand")},
        "order_col_name": order_column,
        "order_col_letter": order_letter,
        "order_col_exists": order_exists,
    }


def _build_outcome(n_restock: int, total: float, filename: str,
                   changeset: writeback.Changeset | None) -> GuidedOutcome:
    """The plan as >=2 ranked, executable options (first = recommended default)."""
    if n_restock > 0:
        key = changeset.idempotency_key if changeset is not None else "n/a"
        options = [
            ExecutionOption(
                label="Apply the staged order quantities to the planilla",
                summary=f"Write the recommended order column into {filename} for {n_restock} SKU(s) "
                        "after your approval - atomic, backed up, rollback-able.",
                score=3.0, recommended=True,
                action=f"follow apply_howto.md in the deliverables: re-stage (same key={key}) then "
                       "writeback.approve + writeback.apply (reversible)",
                tradeoffs="lowest touch; the planilla stays the single source of truth",
            ),
            ExecutionOption(
                label="Export the plan for review",
                summary="Hand off the replenishment plan without touching the client's file.",
                score=1.0,
                action="export the replenishment plan (no write-back)",
                tradeoffs="zero risk; manual follow-up",
            ),
        ]
        summary = f"{n_restock} SKU(s) below target ({total:,.0f} units short): choose how to act."
    else:
        options = [
            ExecutionOption(
                label="Hold - every SKU is above target",
                summary="No replenishment is needed now; nothing to write.",
                score=3.0, recommended=True,
                action="monitor; no write-back needed", tradeoffs="no cost",
            ),
            ExecutionOption(
                label="Tighten the target",
                summary="Lower the cover / order-up-to factor to release working capital.",
                score=2.0,
                action="re-run with a lower cover_periods / order_up_to_factor",
                tradeoffs="frees cash; less buffer",
            ),
        ]
        summary = "All SKUs above target: choose how to proceed."
    return as_options(summary, options)


def run(
    payload: dict,
    *,
    cover_periods: float = 8.0,
    order_up_to_factor: float = 2.0,
    idempotency_key: str | None = None,
) -> ExcelReplenishmentReport:
    """Plan the restock and stage the write-back as a dry-run changeset.

    ``idempotency_key`` defaults to a hash of the staged content, so a NEW plan
    (different quantities / different file state) never collides with a previous
    apply, while re-staging the identical plan stays idempotent.
    """
    if cover_periods <= 0 or order_up_to_factor <= 0:
        raise ValueError("cover_periods and order_up_to_factor must be > 0")
    mode: str = payload["mode"]
    lines: list[ReplenishmentLine] = []
    restock: dict[str, tuple[int, float]] = {}  # sku -> (sheet row, qty)
    n_unplanned = 0
    for row in payload["rows"]:
        # A blank planning signal must never coalesce to 0 (that would silently
        # never replenish - or, with negative stock, order against a ROP of 0).
        if mode == "demand-cover":
            if row.demand_per_period is None:
                n_unplanned += 1
                continue
            target = row.demand_per_period * cover_periods
            qty = max(0.0, round(target - row.on_hand, 1))
        else:
            if row.reorder_point is None:
                n_unplanned += 1
                continue
            target = row.reorder_point * order_up_to_factor
            qty = max(0.0, round(target - row.on_hand, 1)) if row.on_hand < row.reorder_point else 0.0
        lines.append(ReplenishmentLine(row.sku, row.on_hand, target, qty))
        if qty > 0:
            restock[row.sku] = (row.row, qty)

    changeset = None
    if restock:
        store = payload["store"]
        letter = payload["order_col_letter"]
        cells: dict[str, object] = {}
        if not payload["order_col_exists"]:
            cells[f"{letter}{payload['header_row']}"] = payload["order_col_name"]
        for _sku, (r, qty) in restock.items():
            cells[f"{letter}{r}"] = qty
        # GUARD cells: stage the plan's INPUTS as no-op changes (before == after,
        # values as the store reads them) so the connector's commit-time drift
        # check also refuses when stock/ROP/demand changed after staging - the
        # staleness window between staging and the human apply is unbounded.
        current = store.read(payload["sheet"])
        input_cols = payload["input_cols"]
        signal_letter = input_cols["demand"] if mode == "demand-cover" else input_cols["rop"]
        for _sku, (r, _qty) in restock.items():
            for col_letter in (input_cols["stock"], signal_letter):
                guard_cell = f"{col_letter}{r}"
                cells.setdefault(guard_cell, current.get(guard_cell))
        if idempotency_key is None:
            content = repr((payload["filename"], payload["sheet"], sorted(cells.items(), key=lambda kv: kv[0])))
            idempotency_key = "excel-replenish-" + hashlib.sha256(content.encode("utf-8")).hexdigest()[:12]
        changeset = writeback.stage(
            store, f"excel:{payload['filename']}", {payload["sheet"]: cells},
            risk_tier=writeback.TIER_REVERSIBLE, idempotency_key=idempotency_key,
            reason=f"replenish {len(restock)} SKU(s) to target ({mode})",
        )

    flat_restock = {sku: qty for sku, (_r, qty) in restock.items()}
    total = sum(flat_restock.values())
    excluded = ""
    if n_unplanned or payload["n_skipped_rows"]:
        parts = []
        if n_unplanned:
            signal = "demand" if mode == "demand-cover" else "reorder point"
            parts.append(f"{n_unplanned} SKU(s) NOT planned (blank {signal})")
        if payload["n_skipped_rows"]:
            parts.append(f"{payload['n_skipped_rows']} row(s) skipped (non-numeric stock)")
        excluded = " " + "; ".join(parts) + "."
    summary = (
        f"Read {len(lines)} plannable SKU(s) from {payload['filename']} ({payload['sheet']}, {mode}); "
        + (f"{len(flat_restock)} below target ({total:,.0f} units short), staged as a dry-run."
           if flat_restock else "all above target - nothing to write.")
        + excluded
    )
    return ExcelReplenishmentReport(
        filename=payload["filename"],
        sheet=payload["sheet"],
        mode=mode,
        lines=tuple(lines),
        restock=flat_restock,
        n_skus=len(lines),
        n_restock=len(flat_restock),
        n_unplanned=n_unplanned,
        n_skipped_rows=payload["n_skipped_rows"],
        total_restock=total,
        cover_periods=cover_periods,
        order_up_to_factor=order_up_to_factor,
        changeset=changeset,
        outcome=_build_outcome(len(flat_restock), total, payload["filename"], changeset),
        summary=summary,
    )


def verify(report: ExcelReplenishmentReport) -> list[str]:
    """QA gate: protected outcome, real rows, and staging consistent with the plan."""
    issues = list(verify_guided(report.outcome))
    if report.n_skus == 0:
        issues.append(
            "no plannable SKU rows (every row lacked its planning signal) - "
            "fill the reorder-point/demand column or pass the right *_column params"
        )
    if any(ln.restock_qty < 0 for ln in report.lines):
        issues.append("negative restock quantity in the plan")
    if report.restock and report.changeset is None:
        issues.append("restock planned but no changeset was staged")
    if not report.restock and report.changeset is not None:
        issues.append("a changeset was staged with nothing to restock")
    return issues


_HOWTO_TEMPLATE = """# How to apply the staged replenishment

The plan below was STAGED as a dry-run - nothing has been written to
`{filename}`. To apply it (atomic write, byte-exact backup, rollback-able),
run the operator CLI from the repo root:

```
python examples/apply_replenishment.py --file "<path to {filename}>"
```

It re-plans from the CURRENT file, shows the exact before/after of every cell,
asks for your confirmation, applies, and prints the rollback command. Undo any
time with `--rollback <key>`.

Programmatic alternative (same safety plane):

```python
from jobs import excel_replenishment_job as job
from src import writeback

payload = job.prepare(r"<path to {filename}>", {{}})   # same params as the original run
report = job.run(payload)                              # re-stages; same content => same key
approval = writeback.approve(report.changeset, "your-name")   # 900s TTL
writeback.apply(payload["store"], report.changeset, approval=approval)
# Undo later: payload["store"].rollback(report.changeset.idempotency_key)
```

If the planilla changed since this plan was made, the re-run produces a NEW plan
(different key) from the current file - that is the drift protection working:
you always approve exactly what will be written, computed from current data.

Staged plan (key `{key}`): {n} SKU(s), {total:,.0f} units.
"""


def write_operational(report: ExcelReplenishmentReport, out_dir: str | Path,
                      client: str = "Client") -> dict[str, Path]:
    """Machine-readable deliverable + the exact apply recipe for the staged plan."""
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    rows = [
        {
            "sku": ln.sku,
            "on_hand": round(ln.on_hand, 2),
            "target": round(ln.target, 2),
            "restock_qty": round(ln.restock_qty, 2),
        }
        for ln in report.lines
    ]
    written = {"csv": write_summary_csv(rows, d / "excel_replenishment.csv")}
    if report.changeset is not None:
        howto = d / "apply_howto.md"
        howto.write_text(
            _HOWTO_TEMPLATE.format(
                filename=report.filename, key=report.changeset.idempotency_key,
                n=report.n_restock, total=report.total_restock,
            ),
            encoding="utf-8",
        )
        written["apply_howto"] = howto
    return written


def build_deck(
    report: ExcelReplenishmentReport,
    *,
    client: str = "Client",
    prepared: str = "",
    citations: tuple[str, ...] = (),
    confidence: float = 0.85,
) -> Deliverable:
    """Compose the study: what is short, by how much, and how to act on the planilla."""
    short = [ln for ln in report.lines if ln.restock_qty > 0]
    findings = [
        Finding(
            f"{report.n_restock} SKU(s) below target",
            f"{report.total_restock:,.0f} units short across {report.n_skus} plannable SKU(s) read "
            f"from {report.filename} ({report.sheet}, {report.mode} mode).",
            impact="replenish to avoid stockouts on the thin SKUs",
        )
    ]
    if short:
        worst = max(short, key=lambda ln: ln.restock_qty)
        findings.append(
            Finding(
                f"Thinnest SKU: {worst.sku}",
                f"{worst.on_hand:.0f} on hand vs a {worst.target:.0f} target (+{worst.restock_qty:.0f} needed).",
                impact="prioritize this replenishment line",
            )
        )
    if report.n_unplanned or report.n_skipped_rows:
        signal = "demand" if report.mode == "demand-cover" else "reorder point"
        findings.append(
            Finding(
                "Rows the plan could NOT cover",
                f"{report.n_unplanned} SKU(s) with a blank {signal} and {report.n_skipped_rows} row(s) "
                "with non-numeric stock were excluded - they are invisible to this plan.",
                impact=f"fill the {signal}/stock cells so every SKU is protected",
            )
        )
    kpis = (
        Kpi("SKUs planned", str(report.n_skus), rationale=f"From {report.filename}"),
        Kpi("SKUs to replenish", str(report.n_restock), target="0", rationale="Below target"),
        Kpi("Units short", f"{report.total_restock:,.0f}", target="0", rationale="Restock to reach target"),
        Kpi(
            "Target rule",
            f"{report.cover_periods:.0f} periods of demand" if report.mode == "demand-cover"
            else f"{report.order_up_to_factor:.1f} x reorder point",
            rationale="How the target was set",
        ),
    )
    data_sources = (
        DataSource("Stock / reorder data", f"Client planilla {report.filename} ({report.sheet})", "on run"),
        DataSource("Order-quantity write-back", "Staged via src/connectors/excel.py (safe-staging plane)",
                   "on apply"),
    )
    recommendations = [
        "Approve and apply the staged order column so the planilla itself carries the plan.",
        "Keep the reorder/demand columns current - the plan is only as good as the planilla.",
    ]
    return Deliverable(
        title="Excel Replenishment",
        client=client,
        summary=report.summary,
        findings=tuple(findings),
        kpis=kpis,
        data_sources=data_sources,
        recommendations=tuple(recommendations),
        citations=tuple(citations),
        confidence=confidence,
        residual="Applying the staged order quantities writes to the client's file through the "
                 "Excel connector's safe-staging plane (drift check - covering the plan's inputs "
                 "via guard cells - backup, atomic write, rollback); a human approves before "
                 "anything is committed.",
        prepared=prepared,
    )
