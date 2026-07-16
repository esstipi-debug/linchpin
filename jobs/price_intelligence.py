"""Price intelligence one-shot playbook (Linchpin 3.0 PR-13, plan sections
6.9 / 3.1 "El Atajo de Revenue"): a client-supplied refs CSV (``product_id``
+ one-or-more competitor URLs per product) -> acquire (PR-11's extraction
cascade against each URL) -> sanity (PR-12's quarantine gate) -> a
price-position deliverable, following the exact same
``intake -> run -> qa -> deliver`` playbook shape every other ``jobs/*.py``
module uses -- no scheduler, no event bus (plan S3.1: the one-shot mode
sidesteps the matching/continuous-monitoring bottleneck entirely).

One-shot semantics (plan S6.5's last line): the client's URL <-> SKU mapping
in the refs CSV IS the match -- ``match_confidence=1.0``, method "human"
in spirit (client-supplied), no GTIN/fuzzy/probabilistic pipeline (that is
PR-14's ``match/`` package, for the discovery-assisted mode).

Acquisition scope for this PR: **L1 only** (structured-data extraction
against a fetched PDP page) -- L0 marketplace APIs (MELI) and L2 watcher
integration are PR-15. A refs row may supply either a live ``competitor_url``
(fetched via ``src.pricing_intel.acquire.pdp_fetcher``, gated by
``require_approved_site``/``CircuitBreaker`` exactly like a continuous
fetcher would be) or a pre-fetched local ``html_path`` snapshot (no network
at all -- still L1 acquisition, just an already-acquired copy of the same
kind of page; useful for a client who already pulled PDP snapshots, and what
this module's own tests/CLI demo use to stay fully offline and deterministic).
A ref whose ``competitor_url`` is not a URL at all (a bare marketplace id,
e.g. an MLA/ASIN) is honestly reported as skipped ("id_ref_requires_l0_api")
rather than guessed at -- PR-15 adds the L0 API that resolves those.

No silent caps (golden rule 14): every ref that does not end up as an
accepted, ledger-appended offer is recorded in ``PriceIntelReport.rows``
with a machine-readable ``reason`` -- site not approved, circuit open,
fetch/extraction failed, FX rate unavailable, sanity-discarded, or
sanity-quarantined. ``jobs.qa.verify_price_intel`` and this module's own
deliverable both surface those rows; nothing is ever just dropped.
"""

from __future__ import annotations

import statistics
from collections import Counter
from dataclasses import dataclass, replace
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlparse

import httpx
import pandas as pd

from scm_agent.citation_gate import filter_citations
from scm_agent.events import Event, EventLedger
from scm_agent.knowledge import KnowledgeBase
from src import i18n
from src.deliverable import DEFAULT_BRANDING, Branding, DataSource, Deliverable, Finding, Kpi
from src.export import write_summary_csv
from src.pricing_intel.acquire.base import CircuitBreaker
from src.pricing_intel.acquire.l1 import AcquisitionSkipped, acquire_l1_offer
from src.pricing_intel.ledger import PriceLedger, default_ledger
from src.pricing_intel.models import CompetitorOffer
from src.pricing_intel.normalize import PriceNormalizationError, convert_to_base_currency
from src.pricing_intel.sanity import (
    SanityStatus,
    check_basic_validity,
    check_intraday_delta,
    check_mad_outlier,
    check_staleness,
    to_competitor_offer,
)

TOOL_KEY = "price_intelligence"

DEFAULT_SLA_HOURS = 48.0
MAD_WINDOW_DAYS = 30
DEFAULT_BREAKER_KWARGS: dict = {}

_PRODUCT_COLS = ("product_id", "sku", "SKU", "Product", "product")
_URL_COLS = ("competitor_url", "url", "competitor_id", "competitor_ref", "link", "URL")
_SITE_COLS = ("competitor_site", "site", "domain")
_OUR_PRICE_COLS = ("our_price", "client_price", "current_price", "list_price")
_CURRENCY_COLS = ("currency", "price_currency", "moneda")
_HTML_PATH_COLS = ("html_path", "html_file", "snapshot_path")


@dataclass(frozen=True)
class PriceIntelRef:
    """One (product, competitor) pairing from the client's refs CSV -- a
    CONFIRMED match by construction (plan S6.5's one-shot exemption)."""

    product_id: str
    competitor_url: str
    site: str | None  # None => competitor_url is not a resolvable URL (an id-only ref)
    our_price: Decimal | None
    currency_hint: str | None
    html_path: str | None


@dataclass(frozen=True)
class RowOutcome:
    """What happened to one :class:`PriceIntelRef` this run -- always
    recorded, never silently dropped (golden rule 14)."""

    product_id: str
    site: str | None
    competitor_url: str
    status: str  # "accepted" | "quarantined" | "discarded" | "skipped"
    reason: str
    offer: CompetitorOffer | None = None


@dataclass(frozen=True)
class PriceIntelReport:
    n_products: int
    n_products_covered: int
    coverage_pct: float
    offers: tuple[CompetitorOffer, ...]
    our_prices: dict[str, Decimal]
    rows: tuple[RowOutcome, ...]
    quarantine_rate: float
    avg_freshness_hours: float
    sla_hours: float
    tier_mix: dict[str, int]
    stale_events: tuple[Event, ...]
    now: datetime
    summary: str

    @property
    def quarantined(self) -> tuple[RowOutcome, ...]:
        return tuple(r for r in self.rows if r.status == "quarantined")

    @property
    def discarded(self) -> tuple[RowOutcome, ...]:
        return tuple(r for r in self.rows if r.status == "discarded")

    @property
    def skipped(self) -> tuple[RowOutcome, ...]:
        return tuple(r for r in self.rows if r.status == "skipped")


def _pick_column(df: pd.DataFrame, override: object, candidates: tuple[str, ...]) -> str | None:
    if override:
        return str(override) if str(override) in df.columns else None
    return next((c for c in candidates if c in df.columns), None)


def _derive_site(url: str) -> str | None:
    """The bare, normalized domain for ``url`` -- ``None`` when ``url`` is
    not actually a URL (a bare marketplace id), matching
    ``models.CompetitorOffer.site``'s "no scheme, no spaces" contract."""
    parsed = urlparse(url.strip())
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        return None
    host = parsed.netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def prepare_records(df: pd.DataFrame, params: dict | None = None, *, base_dir: str | Path | None = None) -> dict:
    """Sniff the refs columns and build one :class:`PriceIntelRef` per row."""
    params = params or {}
    product = _pick_column(df, params.get("product_col"), _PRODUCT_COLS)
    url = _pick_column(df, params.get("url_col"), _URL_COLS)
    missing = [n for n, c in (("product_id", product), ("competitor_url", url)) if c is None]
    if missing:
        cols = list(df.columns)[:10]
        raise ValueError(f"could not find {', '.join(missing)}; pass them in params (columns seen: {cols})")

    site_col = _pick_column(df, params.get("site_col"), _SITE_COLS)
    price_col = _pick_column(df, params.get("our_price_col"), _OUR_PRICE_COLS)
    currency_col = _pick_column(df, params.get("currency_col"), _CURRENCY_COLS)
    html_col = _pick_column(df, params.get("html_path_col"), _HTML_PATH_COLS)
    root = Path(base_dir) if base_dir is not None else Path(".")

    refs: list[PriceIntelRef] = []
    our_prices: dict[str, Decimal] = {}
    for _, row in df.iterrows():
        pid = str(row[product]).strip()
        competitor_url = str(row[url]).strip()
        site = str(row[site_col]).strip().lower() if site_col and pd.notna(row[site_col]) else _derive_site(competitor_url)
        currency_hint = str(row[currency_col]).strip().upper() if currency_col and pd.notna(row[currency_col]) else None
        html_path = None
        if html_col and pd.notna(row[html_col]) and str(row[html_col]).strip():
            candidate_path = Path(str(row[html_col]).strip())
            html_path = str(candidate_path if candidate_path.is_absolute() else root / candidate_path)
        our_price = None
        if price_col and pd.notna(row[price_col]):
            our_price = Decimal(str(row[price_col]))
            our_prices.setdefault(pid, our_price)
        refs.append(PriceIntelRef(
            product_id=pid, competitor_url=competitor_url, site=site,
            our_price=our_price, currency_hint=currency_hint, html_path=html_path,
        ))

    return {
        "refs": refs,
        "our_prices": our_prices,
        "sla_hours": float(params.get("sla_hours", DEFAULT_SLA_HOURS)),
    }


def prepare(data_path: str, params: dict | None = None) -> dict:
    """Read a refs CSV and build the price-intel payload. Own reader
    (pandas directly), NOT ``jobs/intake.py`` -- this is a SKU<->competitor
    mapping, not a demand-history file (see this module's own docstring)."""
    return prepare_records(pd.read_csv(data_path), params, base_dir=Path(data_path).resolve().parent)


# -- acquisition ---------------------------------------------------------------


def _acquire_one(
    ref: PriceIntelRef,
    *,
    ledger: PriceLedger,
    event_ledger: EventLedger | None,
    site_configs: dict,
    breakers: dict,
    client: httpx.Client,
    sites_config_dir: str | Path | None,
    now: datetime,
) -> RowOutcome:
    """Acquire one ref's current L1 price via the shared
    ``src.pricing_intel.acquire.l1.acquire_l1_offer`` prefix (final whole-
    branch review, Finding 2 -- this function used to carry the gate/tier/
    breaker/fetch/classify/extract sequence inline, near-verbatim with
    ``jobs.price_watch._check_one_pair``'s own copy; both now call the same
    shared helper). Everything from here down is this function's OWN
    divergent tail: the multi-check sanity gate (basic validity / intraday
    delta / MAD outlier) plus FX normalization -- ``jobs.price_watch``
    converges on ``accept_observation`` instead, a deliberately different
    tail (see that module's own docstring)."""
    base = dict(product_id=ref.product_id, site=ref.site, competitor_url=ref.competitor_url)

    acquired = acquire_l1_offer(
        site=ref.site, competitor_ref=ref.competitor_url, matched_product_id=ref.product_id,
        match_confidence=1.0, client=client, now=now, site_configs=site_configs, breakers=breakers,
        sites_config_dir=sites_config_dir, event_ledger=event_ledger, html_path=ref.html_path,
        currency_hint=ref.currency_hint, breaker_kwargs=DEFAULT_BREAKER_KWARGS,
    )
    if isinstance(acquired, AcquisitionSkipped):
        if acquired.reason == "extraction_failed":
            # This module's OWN extraction_failed event shape (source/payload
            # keys) -- the shared prefix deliberately never emits this event
            # itself (see acquire_l1_offer's own docstring).
            if event_ledger is not None:
                event_ledger.emit(Event(
                    type="extraction_failed", severity="warning", source="jobs.price_intelligence",
                    dedup_key=f"extraction_failed:{ref.site}:{ref.competitor_url}:{now.isoformat()}",
                    sku=ref.product_id,
                    payload={
                        "site": ref.site, "competitor_url": ref.competitor_url,
                        "attempts": list(acquired.extraction_attempts or ()),
                    },
                    ts=now,
                ))
        return RowOutcome(**base, status="skipped", reason=acquired.reason)

    candidate = acquired.candidate
    verdict = check_basic_validity(candidate, ledger=event_ledger)
    if verdict.status == SanityStatus.DISCARD:
        return RowOutcome(**base, status="discarded", reason=verdict.reason)

    try:
        price_normalized = convert_to_base_currency(candidate.price, candidate.currency)
    except PriceNormalizationError:
        return RowOutcome(**base, status="skipped", reason="fx_rate_unavailable")
    candidate = replace(candidate, price_normalized=price_normalized)

    previous = ledger.latest_by_sku(ref.site, ref.competitor_url)
    previous_price = previous.offer.price_normalized if previous is not None else None
    delta_verdict = check_intraday_delta(candidate, previous_price, ledger=event_ledger)
    if delta_verdict.status == SanityStatus.QUARANTINE:
        return RowOutcome(**base, status="quarantined", reason=delta_verdict.reason)

    window_start = now - timedelta(days=MAD_WINDOW_DAYS)
    trailing_window = [
        r.offer.price_normalized for r in ledger.history_for_sku(ref.site, ref.competitor_url)
        if r.offer.observed_at >= window_start
    ]
    mad_verdict = check_mad_outlier(candidate, trailing_window, ledger=event_ledger)
    if mad_verdict.status == SanityStatus.QUARANTINE:
        return RowOutcome(**base, status="quarantined", reason=mad_verdict.reason)

    return RowOutcome(**base, status="accepted", reason="ok", offer=to_competitor_offer(candidate))


def run(
    payload: dict,
    *,
    ledger: PriceLedger | None = None,
    event_ledger: EventLedger | None = None,
    http_client: httpx.Client | None = None,
    sites_config_dir: str | Path | None = None,
    now: datetime | None = None,
) -> PriceIntelReport:
    """Acquire -> sanity-gate -> append every ref, then roll up the report."""
    ledger = ledger or default_ledger()
    now = now or datetime.now(timezone.utc)
    refs: list[PriceIntelRef] = payload["refs"]
    sla_hours: float = payload["sla_hours"]
    our_prices: dict[str, Decimal] = dict(payload["our_prices"])

    site_configs: dict[str, object] = {}
    breakers: dict[str, CircuitBreaker] = {}
    owns_client = http_client is None
    client = http_client or httpx.Client()
    try:
        rows = [
            _acquire_one(
                ref, ledger=ledger, event_ledger=event_ledger, site_configs=site_configs,
                breakers=breakers, client=client, sites_config_dir=sites_config_dir, now=now,
            )
            for ref in refs
        ]
    finally:
        if owns_client:
            client.close()

    accepted_offers = [r.offer for r in rows if r.status == "accepted" and r.offer is not None]
    if accepted_offers:
        ledger.append(accepted_offers, now=now)

    stale_events: list[Event] = []
    for ref in refs:
        if ref.site is None or isinstance(site_configs.get(ref.site), Exception):
            continue
        prior = ledger.latest_by_sku(ref.site, ref.competitor_url)
        if prior is None:
            continue
        ev = check_staleness(
            site=ref.site, competitor_sku_ref=ref.competitor_url, matched_product_id=ref.product_id,
            last_observed_at=prior.offer.observed_at, sla_hours=sla_hours, now=now, ledger=event_ledger,
        )
        if ev is not None:
            stale_events.append(ev)

    n_products = len({r.product_id for r in refs}) if refs else 0
    covered_products = {o.matched_product_id for o in accepted_offers if o.matched_product_id}
    n_products_covered = len(covered_products)
    coverage_pct = (n_products_covered / n_products) if n_products else 0.0
    quarantined_count = sum(1 for r in rows if r.status == "quarantined")
    quarantine_rate = (quarantined_count / len(rows)) if rows else 0.0
    if accepted_offers:
        freshness_hours = [(now - o.observed_at).total_seconds() / 3600.0 for o in accepted_offers]
        avg_freshness_hours = statistics.fmean(freshness_hours)
    else:
        avg_freshness_hours = 0.0
    tier_mix = dict(Counter(o.acquisition_tier for o in accepted_offers))

    summary = (
        f"Price position across {n_products} product(s): {n_products_covered} "
        f"({coverage_pct * 100:.0f}%) have >=1 confirmed competitor observation. "
        f"{quarantined_count} quarantined, {sum(1 for r in rows if r.status == 'discarded')} discarded, "
        f"{sum(1 for r in rows if r.status == 'skipped')} skipped -- see Fuentes for every row."
    )

    return PriceIntelReport(
        n_products=n_products, n_products_covered=n_products_covered, coverage_pct=coverage_pct,
        offers=tuple(accepted_offers), our_prices=our_prices, rows=tuple(rows),
        quarantine_rate=quarantine_rate, avg_freshness_hours=avg_freshness_hours, sla_hours=sla_hours,
        tier_mix=tier_mix, stale_events=tuple(stale_events), now=now, summary=summary,
    )


# -- deliverable: price_position_matrix.xlsx + ledger_export.csv --------------


def _position_index(our_price: Decimal | None, competitor_prices: list[Decimal]) -> Decimal | None:
    """Our price / average competitor price -- <1 means we're cheaper, >1
    means we're pricier. ``None`` when either side is unknown (never a
    fabricated 1.0)."""
    if our_price is None or not competitor_prices:
        return None
    avg = sum(competitor_prices) / Decimal(len(competitor_prices))
    if avg == 0:
        return None
    return our_price / avg


def write_operational(report: PriceIntelReport, out_dir: str | Path, client: str = "Client") -> dict[str, Path]:
    """The two raw/tabular deliverables: ``price_position_matrix.xlsx`` and
    ``ledger_export.csv`` (plan section 6.9 item 3). Quarantined/discarded/
    skipped rows get their OWN sheet -- never silently dropped (golden rule 14)."""
    from openpyxl import Workbook

    from src.sanitize import defuse_formula

    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)

    by_product: dict[str, list[CompetitorOffer]] = {}
    for offer in report.offers:
        if offer.matched_product_id:
            by_product.setdefault(offer.matched_product_id, []).append(offer)

    wb = Workbook()
    ws = wb.active
    ws.title = "Position Matrix"
    ws.append(["product_id", "our_price", "competitor_site", "competitor_price_normalized",
               "position_index", "acquisition_tier", "extractor", "confidence", "observed_at"])
    for product_id in sorted(by_product):
        offers = by_product[product_id]
        our_price = report.our_prices.get(product_id)
        idx = _position_index(our_price, [o.price_normalized for o in offers])
        for offer in offers:
            ws.append([
                defuse_formula(product_id), float(our_price) if our_price is not None else None,
                defuse_formula(offer.site), float(offer.price_normalized),
                float(idx) if idx is not None else None, offer.acquisition_tier,
                defuse_formula(offer.extractor), offer.extraction_confidence,
                offer.observed_at.isoformat(),
            ])

    quarantine_ws = wb.create_sheet("Quarantine & Discards")
    quarantine_ws.append(["product_id", "site", "competitor_url", "status", "reason"])
    for row in report.rows:
        if row.status in ("quarantined", "discarded"):
            quarantine_ws.append([
                defuse_formula(row.product_id), defuse_formula(row.site or ""),
                defuse_formula(row.competitor_url), row.status, row.reason,
            ])

    skipped_ws = wb.create_sheet("Skipped")
    skipped_ws.append(["product_id", "site", "competitor_url", "reason"])
    for row in report.rows:
        if row.status == "skipped":
            skipped_ws.append([
                defuse_formula(row.product_id), defuse_formula(row.site or ""),
                defuse_formula(row.competitor_url), row.reason,
            ])

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["metric", "value"])
    summary_ws.append(["client", client])
    summary_ws.append(["n_products", report.n_products])
    summary_ws.append(["n_products_covered", report.n_products_covered])
    summary_ws.append(["coverage_pct", round(report.coverage_pct * 100, 1)])
    summary_ws.append(["quarantine_rate_pct", round(report.quarantine_rate * 100, 1)])
    summary_ws.append(["avg_freshness_hours", round(report.avg_freshness_hours, 2)])
    summary_ws.append(["sla_hours", report.sla_hours])
    for tier, count in sorted(report.tier_mix.items()):
        summary_ws.append([f"observations_tier_{tier}", count])

    matrix_path = d / "price_position_matrix.xlsx"
    wb.save(matrix_path)

    ledger_rows = [
        {
            "observed_at": o.observed_at.isoformat(), "site": o.site,
            "competitor_sku_ref": o.competitor_sku_ref, "matched_product_id": o.matched_product_id,
            "match_confidence": o.match_confidence, "price": str(o.price), "currency": o.currency,
            "price_normalized": str(o.price_normalized), "availability": o.availability,
            "promo_flag": o.promo_flag, "list_price": str(o.list_price) if o.list_price is not None else "",
            "acquisition_tier": o.acquisition_tier, "extractor": o.extractor,
            "extractor_version": o.extractor_version, "extraction_confidence": o.extraction_confidence,
        }
        for o in report.offers
    ]
    ledger_path = (
        write_summary_csv(ledger_rows, d / "ledger_export.csv") if ledger_rows
        else write_summary_csv([{"note": "no accepted observations this run"}], d / "ledger_export.csv")
    )

    return {"matrix": matrix_path, "ledger_csv": ledger_path}


def _render_fuentes_section(report: PriceIntelReport, lang: str) -> str:
    """Golden rule 7's per-datum acquisition-tier table -- ASCII, always
    present, independent of the L3 book-citation gate (that lives in the
    Deliverable's own ``citations``/methodology section)."""
    L = lambda key: i18n.label(key, lang)  # noqa: E731
    out = [f"## {L('hdr_fuentes')}", "", L("fuentes_intro"), ""]
    if report.offers:
        out += [
            f"| {L('fuentes_col_product')} | {L('fuentes_col_site')} | {L('fuentes_col_tier')} | "
            f"{L('fuentes_col_extractor')} | {L('fuentes_col_confidence')} | {L('fuentes_col_observed_at')} |",
            "|---|---|---|---|---|---|",
        ]
        for o in report.offers:
            out.append(
                f"| {o.matched_product_id or '-'} | {o.site} | {o.acquisition_tier} | "
                f"{o.extractor} v{o.extractor_version} | {o.extraction_confidence * 100:.0f}% | "
                f"{o.observed_at.isoformat()} |"
            )
        out.append("")
    if report.quarantined or report.discarded:
        out += [f"### {L('fuentes_quarantine_hdr')}", "", L("fuentes_quarantine_intro"), ""]
        for row in (*report.quarantined, *report.discarded):
            out.append(f"- {row.product_id} @ {row.site or row.competitor_url}: {row.status} ({row.reason})")
        out.append("")
    if report.skipped:
        out += [f"### {L('fuentes_skipped_hdr')}", ""]
        for row in report.skipped:
            out.append(f"- {row.product_id} @ {row.site or row.competitor_url}: {row.reason}")
        out.append("")
    return "\n".join(out)


def build_deck(
    report: PriceIntelReport,
    *,
    client: str = "Client",
    prepared: str = "",
    citations: tuple[str, ...] = (),
    confidence: float = 0.85,
    lang: str = "en",  # matches Deliverable's own field default -- see src/deliverable.py's
    # docstring: individual tool decks (this one included) stay "en" unless a caller
    # explicitly opts into Spanish (the CLI does; the registered Tool's deck does not,
    # exactly like every other jobs/<x>_job.py::build_deck).
    branding: Branding | None = None,
) -> Deliverable:
    """Compose the price-position narrative deck. Matches every other job's
    ``build_deck(report, *, client, prepared, citations, confidence)``
    signature; ``lang``/``branding`` are the E4/E6 hooks golden rule 13
    requires of every 3.0 deliverable."""
    findings = [
        Finding(
            "Competitor coverage",
            f"{report.n_products_covered} of {report.n_products} product(s) have at least one "
            f"confirmed, non-quarantined competitor observation ({report.coverage_pct * 100:.0f}%).",
            impact="the rest have no reliable competitor read this cycle",
        ),
    ]
    if report.quarantined or report.discarded:
        findings.append(Finding(
            "Data quality gate",
            f"{len(report.quarantined)} observation(s) quarantined (unconfirmed jump / outlier) and "
            f"{len(report.discarded)} discarded (invalid price/currency) -- never shipped as if trustworthy.",
            impact="see the Fuentes section and the Quarantine & Discards sheet",
        ))
    if report.skipped:
        findings.append(Finding(
            "Skipped refs",
            f"{len(report.skipped)} ref(s) produced no observation this run (site not approved, "
            "circuit open, fetch/extraction failed, or FX rate unavailable).",
            impact="coverage is lower than the refs file alone would suggest",
        ))

    kpis = (
        Kpi("Products in scope", str(report.n_products), rationale="Distinct product_id in the refs file"),
        Kpi("Coverage", f"{report.coverage_pct * 100:.0f}%", target=">=60%",
            rationale="Products with >=1 confirmed competitor observation"),
        Kpi("Quarantine rate", f"{report.quarantine_rate * 100:.0f}%", target="minimize",
            rationale="Share of ref rows flagged by the sanity gate, not shipped as trustworthy"),
        Kpi("Avg. freshness", f"{report.avg_freshness_hours:.1f}h", target=f"<= {report.sla_hours:.0f}h",
            rationale="Age of the accepted observations vs. the stated SLA"),
    )

    data_sources = (
        DataSource("Competitor price/availability", "PDP pages (L1 structured data)", "per run"),
        DataSource("Our own price", "client refs file", "per run"),
    )

    recommendations = (
        "Act on the accepted (non-quarantined) rows only; a quarantined row needs a confirmatory "
        "second read before it should move a price decision.",
        "Investigate the skipped/discarded refs (site approval, extraction) to raise coverage next cycle.",
        "Re-run on a cadence at or under the stated SLA to keep the position matrix within freshness.",
    )

    return Deliverable(
        title="Price Position Diagnostic",
        client=client,
        summary=report.summary,
        findings=tuple(findings),
        kpis=kpis,
        data_sources=data_sources,
        recommendations=recommendations,
        citations=tuple(citations),
        confidence=confidence,
        residual="This is a one-shot read against the refs supplied: it does not discover new "
                 "competitors, re-price anything, or monitor continuously. Acting on a price change, "
                 "approving a re-price, and refreshing the refs file stay on the operator's side.",
        prepared=prepared,
        lang=lang,
        branding=branding if branding is not None else DEFAULT_BRANDING,
    )


# L3 citations that ground the price-intelligence *method* (which books back
# competitive price positioning), not the client's numbers -- so grounding runs
# on this fixed keyword set, NOT the free-text brief, and is deterministic across
# every deliverable. Same defect + fix as jobs/integrated_plan.py (3.0-audit
# finding #7), reproduced against the committed graph: at the old limit=3 with the
# brief fed into grounding, a realistic "benchmark vs Amazon/MercadoLibre" brief
# grounded islanded case/forecast nodes ("Amazon" -> a Cachon/Tang case node,
# "benchmark" -> forecast nodes) that displaced the real price-position anchors
# past the top-3 pool -> ZERO citations; other briefs dragged in off-topic nodes.
# Grounding the keyword set only, over a wider pool with a tight cap, keeps the
# citations on-topic and present regardless of wording. The gate (anchors,
# MAX_HOPS, MIN_CITATIONS) is untouched.
_CITATION_KEYWORDS = (
    "price competition", "competitor pricing", "price position", "price positioning",
    "market pricing", "price intelligence",
)
_CANDIDATE_POOL = 6   # candidates grounded and offered to the strict gate
_MAX_CITATIONS = 3    # kept, on-topic survivors ultimately shown (tight display)


def gated_citations(
    brief: str = "",
    *,
    kb: KnowledgeBase | None = None,
    limit: int = _CANDIDATE_POOL,
    max_shown: int = _MAX_CITATIONS,
) -> tuple[str, ...]:
    """The E5-gated L3 citations for this tool (golden rule 7 + the citation
    gate) -- ``scm_agent.citation_gate.filter_citations``, reused verbatim
    (no parallel citation mechanism), the exact pattern
    ``scm_agent/packages.py::_run_step`` already established.

    Grounds on the fixed ``_CITATION_KEYWORDS`` set only -- ``brief`` is
    deliberately NOT fed into the query (see the module comment): these citations
    ground the *method*, and the client's wording carries more grounding weight
    than the keyword set, so incidental tokens surfaced off-topic citations past
    the gate. ``limit`` grounds a wider pool than we display so the strict gate
    can reach past graph fragmentation; ``[:max_shown]`` shows a tight set.
    ``brief`` is retained in the signature for call-site stability."""
    kb = kb or KnowledgeBase()
    candidates = kb.ground_citations_detailed(_CITATION_KEYWORDS, "", limit=limit)
    return filter_citations(kb, TOOL_KEY, candidates).kept[:max_shown]


def write_report_md(deliverable: Deliverable, report: PriceIntelReport, out_dir: str | Path, lang: str) -> Path:
    """``report.md`` (not the generic ``deliverable.md`` name every other
    job's ``write_all`` produces) -- the standard sectioned Markdown PLUS
    the golden-rule-7 Fuentes section (per-datum acquisition tier)."""
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    md = deliverable.to_markdown().rstrip() + "\n\n" + _render_fuentes_section(report, lang) + "\n"
    path = d / "report.md"
    path.write_text(md, encoding="utf-8")
    return path


def write_deliverable(
    report: PriceIntelReport,
    *,
    out_dir: str | Path,
    client: str = "Client",
    brief: str = "",
    prepared: str = "",
    lang: str = i18n.DEFAULT_LANG,
    branding: Branding | None = None,
    confidence: float = 0.85,
    kb: KnowledgeBase | None = None,
) -> dict[str, Path]:
    """The full 3-file deliverable (plan section 6.9 item 3): the standalone
    one-shot entry point used by the CLI (``examples/run_price_intel.py``)
    and this module's own tests -- always E5-gates its citations."""
    citations = gated_citations(brief, kb=kb)
    deliverable = build_deck(
        report, client=client, prepared=prepared, citations=citations,
        confidence=confidence, lang=lang, branding=branding,
    )
    written = write_operational(report, out_dir, client)
    written["report_md"] = write_report_md(deliverable, report, out_dir, lang)
    return written
