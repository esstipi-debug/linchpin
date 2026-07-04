"""Safe-staging writeback connector for a client's Excel workbook.

Most small/mid clients keep inventory in a planilla, not an ERP — this connector
makes that planilla a first-class, SAFELY writable system of record. It exposes
the same store surface as ``writeback.InMemoryStore`` / the Odoo stores
(``read`` / ``applied_keys`` / ``claim`` / ``release`` / ``commit`` /
``rollback``), so the entire safety plane — risk tiers, signed time-boxed
approvals, idempotency, audit — is reused from ``src.writeback`` unchanged:

    store = ExcelWorkbookStore("planilla_cliente.xlsx")
    cells = store.resolve_row_edits("Stock", "Codigo", {"SKU-001": {"Punto Reorden": 60}})
    cs = writeback.stage(store, "excel:planilla_cliente.xlsx", {"Stock": cells},
                         risk_tier=writeback.TIER_REVERSIBLE, idempotency_key="rop-2026-07-04")
    approval = writeback.approve(cs, "operator")     # human sees the before/after first
    writeback.apply(store, cs, approval=approval)    # backup + atomic write + audit
    store.rollback("rop-2026-07-04")                 # cell-exact undo, any time later

Addressing: an entity is a SHEET NAME and a field is a CELL REFERENCE ("D5"), so
any workbook works with zero schema assumptions; ``resolve_row_edits`` adds the
semantic layer on top (SKU + column header -> cell), auto-detecting the header
row so a client's own layout (title rows, headers not on row 1) needs no config.

Safety properties beyond the shared plane:
- DRIFT CHECK: commit refuses (nothing written) if any staged cell changed on
  disk after staging — a human editing the file between stage and apply can
  never be silently overwritten.
- ATOMIC WRITE: saved to a temp file then ``os.replace``-d, so a crash or a
  locked file can never leave a half-written workbook.
- FILE BACKUP: every commit first copies the original next to it
  (``<name>.<key>-<contenthash>.linchpin-backup<ext>``) — disaster recovery for
  anything a cell-level rollback can't cover (openpyxl round-trips
  values/formulas/styles, but exotic content like charts can degrade; the backup
  preserves the byte-exact original).
- CRASH-WINDOW TRIPWIRE: if the process dies between the file replace and the
  audit record, the write is live but unaudited — and a naive retry would
  re-stage from the mutated file and record a poisoned restore (rollback would
  then "restore" the new value). The backup left orphaned by the crashed attempt
  is the detector: a commit that finds an existing backup for the same key with
  a DIFFERENT content hash refuses (writing nothing) and directs the operator to
  reconcile from that backup first. In that window the backup — not
  ``rollback()`` — is the authoritative copy of the original.
- ``.xlsm`` macros are PRESERVED (``keep_vba``), never executed — running client
  macros needs live Excel (COM) and is out of scope for a file-level connector.
"""

from __future__ import annotations

import hashlib
import os
import re
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src import writeback

# How deep to scan for a client's header row (planillas often stack title/blank
# rows above the real table; 20 covers every layout seen in client files so far).
_HEADER_SCAN_ROWS = 20

_KEY_SAFE_RE = re.compile(r"[^A-Za-z0-9_-]+")


class ExcelWritebackError(Exception):
    """A file-level writeback failed safely (nothing was changed on disk)."""


def _safe_key(idempotency_key: str) -> str:
    """Filesystem-safe, EXACT-key-unique tag for backup filenames.

    The readable part is sanitized (and may collide: "rop 1" vs "rop-1"); the
    appended hash of the full original key makes the tag unique per exact key, so
    the crash-window tripwire can never fire across two different keys.
    """
    cleaned = _KEY_SAFE_RE.sub("-", idempotency_key).strip("-")[:40] or "key"
    return f"{cleaned}-{hashlib.sha256(idempotency_key.encode('utf-8')).hexdigest()[:8]}"


class ExcelWorkbookStore:
    """writeback system-of-record surface over one Excel workbook file."""

    def __init__(self, path: str | Path, *, ledger: object | None = None,
                 backup_dir: str | Path | None = None) -> None:
        self._path = Path(path)
        self._backup_dir = Path(backup_dir) if backup_dir is not None else self._path.parent
        self._audit = writeback.AuditBookkeeping(ledger)

    # ---- workbook I/O ------------------------------------------------------------

    def _load(self):
        if not self._path.exists():
            raise ExcelWritebackError(f"workbook not found: {self._path}")
        keep_vba = self._path.suffix.lower() in (".xlsm", ".xltm")
        return load_workbook(self._path, keep_vba=keep_vba)

    def _atomic_save(self, wb) -> None:
        fd, tmp_name = tempfile.mkstemp(dir=self._path.parent, prefix=".linchpin-",
                                        suffix=self._path.suffix)
        os.close(fd)
        try:
            wb.save(tmp_name)
            os.replace(tmp_name, self._path)
        except PermissionError as exc:
            raise ExcelWritebackError(
                f"cannot write {self._path.name}: the file appears to be open in Excel — "
                "close it and retry (nothing was changed)"
            ) from exc
        finally:
            Path(tmp_name).unlink(missing_ok=True)

    # ---- writeback store surface ---------------------------------------------------

    def read(self, entity_id: str) -> dict:
        """All non-empty cells of sheet ``entity_id`` as {cell_ref: value}.

        Formula cells read as their formula string (the cell's actual content —
        this connector edits content, it does not compute). A missing sheet reads
        as ``{}``, mirroring ``InMemoryStore``'s unknown entity.
        """
        wb = self._load()
        if entity_id not in wb.sheetnames:
            return {}
        ws = wb[entity_id]
        return {
            cell.coordinate: cell.value
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        }

    def applied_keys(self) -> set[str]:
        return self._audit.applied_keys()

    def claim(self, idempotency_key: str, *, now: float | None = None) -> bool:
        return self._audit.claim(idempotency_key, now=now)

    def release(self, idempotency_key: str) -> None:
        self._audit.release(idempotency_key)

    def commit(self, changeset: writeback.Changeset, approved_by: str) -> writeback.AuditEntry:
        wb = self._load()

        # Fail-closed pre-flight over the LIVE file: every target sheet must exist
        # and every staged cell must still hold its staged `before` — a human who
        # edited the planilla after staging is never silently overwritten. Nothing
        # is written unless every check passes (the workbook is only mutated in
        # memory, and only saved once, atomically, at the end).
        for c in changeset.changes:
            if c.entity_id not in wb.sheetnames:
                raise ExcelWritebackError(
                    f"sheet {c.entity_id!r} not found in {self._path.name} "
                    f"(sheets: {', '.join(wb.sheetnames)})"
                )
            current = wb[c.entity_id][c.field].value
            if current != c.before:
                raise ExcelWritebackError(
                    f"{c.entity_id}!{c.field} changed since staging "
                    f"(file now has {current!r}, staged before was {c.before!r}) — "
                    "re-stage from the current file"
                )

        restore = tuple(
            (
                c.entity_id,
                c.field,
                writeback.ABSENT if wb[c.entity_id][c.field].value is None
                else wb[c.entity_id][c.field].value,
            )
            for c in changeset.changes
        )

        # content_hash suffix disambiguates two changesets whose keys sanitize to
        # the same _safe_key (e.g. "rop 1" vs "rop-1") — neither overwrites the
        # other's backup.
        backup = self._backup_dir / self._backup_name(changeset.idempotency_key, changeset.content_hash)

        # Crash-window tripwire: a backup for this key with a DIFFERENT content
        # hash means an earlier attempt backed up and possibly wrote the file but
        # never recorded its audit (process died in that window). Committing now
        # would bake the mutated file in as "original" and poison rollback — the
        # orphaned backup holds the true original, so reconciliation comes first.
        stale = [p for p in self._backups_for_key(changeset.idempotency_key) if p.name != backup.name]
        if stale:
            raise ExcelWritebackError(
                f"an earlier apply attempt for key {changeset.idempotency_key!r} left a backup but "
                f"no audit record — its write may already be live in {self._path.name}. Reconcile "
                f"against {stale[0].name} (the pre-attempt original), then delete that backup to retry"
            )
        self._backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self._path, backup)

        for c in changeset.changes:
            wb[c.entity_id][c.field] = c.after
        try:
            self._atomic_save(wb)
        except BaseException:
            backup.unlink(missing_ok=True)  # nothing changed; a stray backup would mislead
            raise

        entry = writeback.AuditEntry(
            changeset.idempotency_key, changeset.target, approved_by, restore,
        )
        self._audit.record(entry)
        return entry

    def rollback(self, idempotency_key: str) -> None:
        entry = self._audit.get(idempotency_key)
        if entry is None:
            raise KeyError(idempotency_key)
        wb = self._load()
        for entity_id, field, original in entry.restore:
            if entity_id not in wb.sheetnames:
                raise ExcelWritebackError(
                    f"cannot roll back: sheet {entity_id!r} no longer exists in {self._path.name}"
                )
            wb[entity_id][field] = None if original is writeback.ABSENT else original
        self._atomic_save(wb)
        self._audit.forget(idempotency_key)
        # The backup is redundant once the file is restored — and leaving it would
        # false-trigger the crash-window tripwire on a legitimate reuse of this key.
        for p in self._backups_for_key(idempotency_key):
            p.unlink(missing_ok=True)

    def _backup_name(self, idempotency_key: str, content_hash: str) -> str:
        return (
            f"{self._path.stem}.{_safe_key(idempotency_key)}"
            f"-{content_hash[:8]}.linchpin-backup{self._path.suffix}"
        )

    def _backups_for_key(self, idempotency_key: str) -> list[Path]:
        """Existing backups for this key (any content hash), by exact name shape.

        Manual prefix/suffix matching, not ``glob`` — a client filename may contain
        glob metacharacters (``[]``), which would corrupt a pattern silently.
        """
        if not self._backup_dir.exists():
            return []
        prefix = f"{self._path.stem}.{_safe_key(idempotency_key)}-"
        suffix = f".linchpin-backup{self._path.suffix}"
        return sorted(
            p for p in self._backup_dir.iterdir()
            if p.name.startswith(prefix) and p.name.endswith(suffix)
        )

    # ---- semantic staging helper -----------------------------------------------------

    def resolve_row_edits(self, sheet: str, key_column: str,
                          updates: dict[str, dict[str, object]]) -> dict[str, object]:
        """Translate row-keyed edits into cell edits ready for ``writeback.stage``.

        ``{"SKU-002": {"Punto Reorden": 90}}`` -> ``{"D5": 90}``. The header row is
        auto-detected (first row within the scan window containing ``key_column``),
        so a client's own title rows above the table need no configuration.
        """
        wb = self._load()
        if sheet not in wb.sheetnames:
            raise ExcelWritebackError(
                f"sheet {sheet!r} not found in {self._path.name} (sheets: {', '.join(wb.sheetnames)})"
            )
        ws = wb[sheet]

        header_row, headers = None, {}
        for row in ws.iter_rows(min_row=1, max_row=_HEADER_SCAN_ROWS):
            labels = {str(c.value).strip(): c.column_letter for c in row if c.value is not None}
            if key_column in labels:
                header_row, headers = row[0].row, labels
                break
        if header_row is None:
            raise ExcelWritebackError(
                f"no header row containing {key_column!r} found in the first "
                f"{_HEADER_SCAN_ROWS} rows of {sheet!r}"
            )

        key_letter = headers[key_column]
        row_by_key: dict[str, int] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            raw = ws[f"{key_letter}{r}"].value
            if raw is None:
                continue
            key = str(raw).strip()
            if key in row_by_key:
                # Ambiguous target: writing "the" row for this key could hit either.
                # Fail closed — the operator dedupes (or uses cell addressing) first.
                raise ExcelWritebackError(
                    f"duplicate row key {key!r} under column {key_column!r} in {sheet!r} "
                    f"(rows {row_by_key[key]} and {r}) — resolve the duplicate before writing"
                )
            row_by_key[key] = r

        cells: dict[str, object] = {}
        for row_key, fields in updates.items():
            row = row_by_key.get(str(row_key).strip())
            if row is None:
                raise ExcelWritebackError(
                    f"row key {row_key!r} not found under column {key_column!r} in {sheet!r}"
                )
            for column_header, value in fields.items():
                letter = headers.get(column_header)
                if letter is None:
                    raise ExcelWritebackError(
                        f"column {column_header!r} not found in {sheet!r} "
                        f"(headers: {', '.join(sorted(headers))})"
                    )
                cells[f"{letter}{row}"] = value
        return cells
