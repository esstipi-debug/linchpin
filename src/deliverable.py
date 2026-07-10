"""Client-ready deliverable composer (capability gap #1: artifacts that sell).

The engines compute; consultants get paid for *documents*. This turns any tool's
output into a professional, sectioned deliverable:

- executive summary (narrative)
- quantified findings / recommendations
- a KPI table with selection rationale (why each KPI, what it tells the client)
- a data-source map (where every number came from + refresh cadence)
- L3 citations (book/chapter grounding)
- a coverage & handoff block (confidence + "what I could NOT do, what you must do")
  honoring the never-unprotected contract of the Guided Execution Layer.

Renders to Markdown (universal) and XLSX (clients live in Excel). Engine-agnostic;
a `scm_agent` JobResult composes directly via `from_result`. Pure/deterministic:
the prepared date is passed in, never read from the clock, so output is testable.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from src import i18n


@dataclass(frozen=True)
class Finding:
    """A single result the client should act on; `impact` is the quantified $/%/days."""

    title: str
    detail: str
    impact: str = ""


@dataclass(frozen=True)
class Kpi:
    """A metric with its target and the rationale for why it was chosen."""

    name: str
    value: str
    target: str = ""
    rationale: str = ""


@dataclass(frozen=True)
class DataSource:
    """One row of the data-source map: a metric, where it came from, how often it refreshes."""

    field: str
    source: str
    cadence: str = ""


@dataclass(frozen=True)
class Deliverable:
    """A composed client deliverable. Empty sections are omitted on render."""

    title: str
    client: str
    summary: str
    findings: tuple[Finding, ...] = ()
    kpis: tuple[Kpi, ...] = ()
    data_sources: tuple[DataSource, ...] = ()
    recommendations: tuple[str, ...] = ()
    options: tuple = ()   # ranked ExecutionOptions (recommended first) - the action menu
    citations: tuple[str, ...] = ()
    confidence: float | None = None
    residual: str = ""
    prepared: str = ""
    # Structural-scaffolding language (section headers, table columns, sheet
    # names) - see src/i18n.py. Defaults to "en" so the ~37 individual tool
    # decks (built by jobs/<x>_job.py, which never pass this) render exactly
    # as before; only jobs/package_deliverable.py's CONSOLIDATED deck passes
    # its package's lang explicitly. This does NOT translate the content
    # (summary/findings/KPI values etc.) - only the scaffolding around it;
    # see i18n.py's module docstring for the full scope boundary.
    lang: str = "en"

    def to_markdown(self) -> str:
        """Render a professional, sectioned Markdown document (ASCII-only for cp1252 safety)."""
        L = lambda key: i18n.label(key, self.lang)  # noqa: E731
        out: list[str] = [f"# {self.title} - {self.client}"]
        if self.prepared:
            out.append(f"*{L('hdr_prepared_field')} {self.prepared}*")
        out.append("")

        if self.summary:
            out += [f"## {L('hdr_executive_summary')}", "", self.summary, ""]

        if self.findings:
            out += [f"## {L('hdr_key_findings')}", ""]
            for f in self.findings:
                line = f"- **{f.title}** - {f.detail}"
                if f.impact:
                    line += f" _(impact: {f.impact})_"
                out.append(line)
            out.append("")

        if self.recommendations:
            out += [f"## {L('hdr_recommendations')}", ""]
            out += [f"{i}. {r}" for i, r in enumerate(self.recommendations, 1)]
            out.append("")

        if self.options:
            out += [f"## {L('hdr_options')}", "", L("options_intro"), ""]
            for i, o in enumerate(self.options, 1):
                flag = L("recommended_flag") if getattr(o, "recommended", False) else ""
                out.append(f"{i}. **{o.label}**{flag} - {o.summary}")
                if getattr(o, "action", ""):
                    out.append(f"   - {L('action_label')}: {o.action}")
                if getattr(o, "tradeoffs", ""):
                    out.append(f"   - {L('tradeoff_label')}: {o.tradeoffs}")
            out.append("")

        if self.kpis:
            out += [f"## {L('hdr_kpis')}", "",
                    f"| {L('col_kpi')} | {L('col_value')} | {L('col_target')} | {L('col_why_it_matters')} |",
                    "|---|---|---|---|"]
            for k in self.kpis:
                out.append(f"| {k.name} | {k.value} | {k.target or '-'} | {k.rationale or '-'} |")
            out.append("")

        if self.data_sources:
            out += [f"## {L('hdr_data_sources')}", "", L("data_sources_intro"), "",
                    f"| {L('col_metric')} | {L('col_source')} | {L('col_refresh')} |", "|---|---|---|"]
            for d in self.data_sources:
                out.append(f"| {d.field} | {d.source} | {d.cadence or '-'} |")
            out.append("")

        if self.citations:
            out += [f"## {L('hdr_methodology')}", "", L("methodology_intro"), ""]
            out += [f"- {c}" for c in self.citations]
            out.append("")

        # Coverage & handoff is always shown: it states confidence and the residual
        # human action, so no deliverable ever reads as "fully autonomous" when it isn't.
        out += [f"## {L('hdr_coverage_handoff')}", ""]
        if self.confidence is not None:
            out.append(f"{L('hdr_confidence_field')}: **{self.confidence * 100:.0f}%**.")
        out.append(self.residual or L("no_residual"))
        out.append("")

        return "\n".join(out).rstrip() + "\n"

    def to_excel(self, path: str | Path) -> Path:
        """Write the deliverable as a multi-sheet workbook (Summary/KPIs/Findings/Data Sources)."""
        from openpyxl import Workbook

        L = lambda key: i18n.label(key, self.lang)  # noqa: E731
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        ws = wb.active
        ws.title = L("sheet_summary")
        ws.append([L("hdr_title_field"), self.title])
        ws.append([L("hdr_client_field"), self.client])
        if self.prepared:
            ws.append([L("hdr_prepared_field"), self.prepared])
        if self.confidence is not None:
            ws.append([L("hdr_confidence_field"), f"{self.confidence * 100:.0f}%"])
        ws.append([])
        ws.append([L("hdr_executive_summary")])
        ws.append([self.summary])
        if self.residual:
            ws.append([])
            ws.append([L("hdr_coverage_handoff")])
            ws.append([self.residual])

        if self.kpis:
            k = wb.create_sheet(L("sheet_kpis"))
            k.append([L("col_kpi"), L("col_value"), L("col_target"), L("col_why_it_matters")])
            for kpi in self.kpis:
                k.append([kpi.name, kpi.value, kpi.target, kpi.rationale])

        if self.findings:
            fnd = wb.create_sheet(L("sheet_findings"))
            fnd.append([L("col_finding"), L("col_detail"), L("col_impact")])
            for f in self.findings:
                fnd.append([f.title, f.detail, f.impact])

        if self.data_sources:
            d = wb.create_sheet(L("sheet_data_sources"))
            d.append([L("col_metric"), L("col_source"), L("col_refresh")])
            for ds in self.data_sources:
                d.append([ds.field, ds.source, ds.cadence])

        if self.options:
            o = wb.create_sheet(L("sheet_options"))
            o.append([L("col_option"), L("col_recommended"), L("col_summary"),
                      L("action_label"), L("tradeoff_label")])
            for opt in self.options:
                o.append([opt.label, L("yes_flag") if getattr(opt, "recommended", False) else "",
                          opt.summary, getattr(opt, "action", ""), getattr(opt, "tradeoffs", "")])

        if self.citations:
            c = wb.create_sheet(L("sheet_citations"))
            c.append([L("col_source")])
            for cite in self.citations:
                c.append([cite])

        wb.save(out)
        return out

    def write_all(self, out_dir: str | Path) -> dict[str, Path]:
        """Write both the Markdown report and the XLSX workbook; return their paths."""
        d = Path(out_dir)
        d.mkdir(parents=True, exist_ok=True)
        md = d / "deliverable.md"
        md.write_text(self.to_markdown(), encoding="utf-8")
        xlsx = self.to_excel(d / "deliverable.xlsx")
        return {"report": md, "workbook": xlsx}


def from_result(
    result: object,
    *,
    title: str,
    client: str = "Client",
    prepared: str = "",
    findings: tuple[Finding, ...] = (),
    kpis: tuple[Kpi, ...] = (),
    data_sources: tuple[DataSource, ...] = (),
    recommendations: tuple[str, ...] = (),
    residual: str = "",
) -> Deliverable:
    """Compose a Deliverable from a `scm_agent` JobResult plus caller-supplied detail.

    Reads the stable JobResult fields (`summary`, `confidence`, `citations`) without
    depending on orchestrator internals. Findings / KPIs / data-source map come from
    the caller (the tool that knows its own numbers). `residual` defaults to the
    guided outcome's summary when present, so the handoff block is populated.
    """
    summary = getattr(result, "summary", "") or ""
    confidence = getattr(result, "confidence", None)
    citations = tuple(getattr(result, "citations", ()) or ())
    if not residual:
        guided = getattr(result, "guided", None)
        residual = getattr(guided, "summary", "") or ""
    return Deliverable(
        title=title,
        client=client,
        summary=summary,
        findings=findings,
        kpis=kpis,
        data_sources=data_sources,
        recommendations=recommendations,
        citations=citations,
        confidence=confidence,
        residual=residual,
        prepared=prepared,
    )
