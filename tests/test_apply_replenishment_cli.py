"""Tests for examples/apply_replenishment.py — the operator apply/rollback CLI.

Closes the review gap EXR-3 (PR #109): the staged changeset used to be applicable
only from a Python session. The CLI re-stages from the current file (content-derived
keys make the changeset identical when nothing changed), shows the exact before/after,
asks for confirmation, applies through the safe-staging plane with a persistent
SQLite ledger (idempotency + rollback survive across invocations), and can roll back.
"""

from __future__ import annotations

import re

import pytest
from openpyxl import Workbook, load_workbook

from examples.apply_replenishment import main as cli_main
from jobs import excel_replenishment_job as job

SHEET = "Stock Bodega"


def _make_planilla(path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", 42, 50])
    ws.append(["SKU-002", 130, 80])
    wb.save(path)
    return path


@pytest.fixture
def planilla(tmp_path):
    return _make_planilla(tmp_path / "planilla.xlsx")


@pytest.fixture
def ledger(tmp_path):
    return tmp_path / "ledger.sqlite3"


# ---- job-level: persistent ledger via params -----------------------------------------

def test_prepare_wires_ledger_when_ledger_path_given(planilla, ledger):
    p1 = job.prepare(str(planilla), {"ledger_path": str(ledger)})
    r1 = job.run(p1)
    from src import writeback
    writeback.apply(p1["store"], r1.changeset, approval=writeback.approve(r1.changeset, "op"))
    # A brand-new store instance over the same ledger remembers the applied key.
    p2 = job.prepare(str(planilla), {"ledger_path": str(ledger)})
    assert r1.changeset.idempotency_key in p2["store"].applied_keys()


# ---- CLI: review -> confirm -> apply ---------------------------------------------------

def test_cli_yes_applies_and_reports_key(planilla, ledger, capsys):
    rc = cli_main(["--file", str(planilla), "--ledger", str(ledger), "--yes"])
    assert rc == 0
    out = capsys.readouterr().out
    assert "applied" in out.lower()
    assert re.search(r"excel-replenish-[0-9a-f]{12}", out)  # the rollback key is shown
    ws = load_workbook(planilla)[SHEET]
    assert ws["D1"].value == "Pedir (Linchpin)"
    assert ws["D2"].value == 58.0


def test_cli_without_confirmation_writes_nothing(planilla, ledger, capsys, monkeypatch):
    monkeypatch.setattr("builtins.input", lambda *_: "n")
    rc = cli_main(["--file", str(planilla), "--ledger", str(ledger)])
    assert rc == 1
    assert load_workbook(planilla)[SHEET]["D2"].value is None
    assert "aborted" in capsys.readouterr().out.lower()


def test_cli_shows_before_after_prior_to_confirmation(planilla, ledger, capsys, monkeypatch):
    monkeypatch.setattr("builtins.input", lambda *_: "n")
    cli_main(["--file", str(planilla), "--ledger", str(ledger)])
    out = capsys.readouterr().out
    assert "SKU-001" in out
    assert "58" in out  # the planned quantity is visible before approving


def test_cli_noop_when_plan_already_applied(planilla, ledger, capsys):
    cli_main(["--file", str(planilla), "--ledger", str(ledger), "--yes"])
    rc = cli_main(["--file", str(planilla), "--ledger", str(ledger), "--yes"])
    assert rc == 0
    assert "already" in capsys.readouterr().out.lower()  # no-op detected, not re-applied


def test_cli_nothing_to_do_on_healthy_planilla(tmp_path, ledger, capsys):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-1", 500, 10])
    f = tmp_path / "healthy.xlsx"
    wb.save(f)
    rc = cli_main(["--file", str(f), "--ledger", str(ledger), "--yes"])
    assert rc == 0
    assert "nothing" in capsys.readouterr().out.lower()
    assert load_workbook(f)[SHEET]["D1"].value is None


# ---- CLI: rollback across processes (persistent ledger) --------------------------------

def test_cli_rollback_restores_file(planilla, ledger, capsys):
    cli_main(["--file", str(planilla), "--ledger", str(ledger), "--yes"])
    key = re.search(r"excel-replenish-[0-9a-f]{12}", capsys.readouterr().out).group(0)
    rc = cli_main(["--file", str(planilla), "--ledger", str(ledger), "--rollback", key])
    assert rc == 0
    ws = load_workbook(planilla)[SHEET]
    assert ws["D2"].value is None  # order qty cleared
    assert ws["B2"].value == 42    # client data intact


def test_cli_rollback_unknown_key_fails_clearly(planilla, ledger, capsys):
    rc = cli_main(["--file", str(planilla), "--ledger", str(ledger), "--rollback", "no-such-key"])
    assert rc == 1
    assert "unknown" in capsys.readouterr().out.lower()


def test_cli_missing_file_fails_clearly(tmp_path, ledger, capsys):
    rc = cli_main(["--file", str(tmp_path / "ghost.xlsx"), "--ledger", str(ledger), "--yes"])
    assert rc == 1
