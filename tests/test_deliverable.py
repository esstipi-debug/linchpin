"""Tests for the client-ready deliverable composer (capability gap #1)."""
from __future__ import annotations

from dataclasses import dataclass

import pytest
from openpyxl import load_workbook

from src.deliverable import DEFAULT_BRANDING, Branding, DataSource, Deliverable, Finding, Kpi, from_result


def _sample() -> Deliverable:
    return Deliverable(
        title="Inventory Optimization",
        client="Acme DTC",
        summary="Analyzed 320 SKUs; recommended $48,000 inventory investment at 95% service level.",
        findings=(
            Finding("12 SKUs at stockout risk", "Reorder now to hold 95% service", impact="$9,200 sales protected"),
            Finding("18 SKUs overstocked", "Excess vs target", impact="$14,500 cash to release"),
        ),
        kpis=(
            Kpi("Inventory turns", "6.2", "8-12", "Cash efficiency of stock"),
            Kpi("Fill rate", "94%", "95%+", "Service the customer sees"),
        ),
        data_sources=(
            DataSource("Demand history", "shopify_orders.csv", "weekly"),
            DataSource("On-hand stock", "Cin7 export", "daily"),
        ),
        recommendations=("Issue POs for the 12 at-risk SKUs", "Mark down the 18 overstocked SKUs"),
        citations=("Reorder Point - vandeput-...pdf Ch.4  -> src/eoq.py:L12",),
        confidence=0.9,
        residual="Approve the drafted POs and confirm supplier lead times before issuing.",
        prepared="2026-06-23",
    )


def test_markdown_has_all_sections_and_preserves_numbers():
    md = _sample().to_markdown()
    for section in ["# Inventory Optimization - Acme DTC", "## Executive summary",
                    "## Key findings", "## Recommendations", "## KPIs",
                    "## Data sources", "## Methodology & grounding", "## Coverage & handoff"]:
        assert section in md
    assert "$48,000" in md and "94%" in md and "$9,200 sales protected" in md
    assert "Confidence: **90%**" in md
    assert "shopify_orders.csv" in md  # data-source map present


def test_empty_sections_are_omitted_but_coverage_always_shown():
    d = Deliverable(title="Bare", client="C", summary="just a summary")
    md = d.to_markdown()
    assert "## Key findings" not in md
    assert "## KPIs" not in md
    assert "## Data sources" not in md
    assert "## Coverage & handoff" in md  # always present
    assert "ready to use" in md  # default residual when none given


def test_markdown_is_ascii_safe():
    # cp1252 console safety: the report must encode without error.
    _sample().to_markdown().encode("cp1252")


def test_to_excel_writes_expected_sheets(tmp_path):
    path = _sample().to_excel(tmp_path / "d.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    assert {"Summary", "KPIs", "Findings", "Data Sources", "Citations"} <= set(wb.sheetnames)
    assert wb["KPIs"].max_row == 3  # header + 2 KPIs


def test_write_all_emits_report_and_workbook(tmp_path):
    paths = _sample().write_all(tmp_path)
    assert paths["report"].exists() and paths["report"].suffix == ".md"
    assert paths["workbook"].exists() and paths["workbook"].suffix == ".xlsx"


# ---- E6: Branding (name/logo/color) -------------------------------------

def test_branding_requires_a_name():
    with pytest.raises(ValueError, match="name"):
        Branding(name="  ")


def test_branding_none_name_raises_valueerror_not_a_raw_attributeerror():
    # A hand-edited/corrupt profile.json can carry `"branding": {"name": null}`
    # - it must fail the same clean ValueError as every other invalid input,
    # not an unwrapped AttributeError client_profile.load_profile can't catch.
    with pytest.raises(ValueError, match="name"):
        Branding(name=None)


def test_branding_name_made_only_of_zero_width_characters_is_rejected():
    # ZERO WIDTH SPACE/JOINER/WORD JOINER/BOM are invisible but not
    # `str.strip()`-whitespace - a name made only of these would otherwise
    # pass validation and render as a silently blank "Prepared by " footer.
    for invisible in ("​", "‍", "⁠", "﻿", "  ​  "):
        with pytest.raises(ValueError, match="name"):
            Branding(name=invisible)


def test_branding_rejects_a_malformed_hex_color():
    with pytest.raises(ValueError, match="primary_color"):
        Branding(name="Acme", primary_color="blue")
    with pytest.raises(ValueError, match="primary_color"):
        Branding(name="Acme", primary_color="#12")


def test_branding_rejects_a_hex_color_with_a_trailing_newline():
    # re.match's unanchored `$` matches just before a trailing "\n" too -
    # fullmatch (or a \Z anchor) is required to genuinely enforce '#RRGGBB'.
    with pytest.raises(ValueError, match="primary_color"):
        Branding(name="Acme", primary_color="#AABBCC\n")


def test_branding_accepts_a_well_formed_hex_color():
    b = Branding(name="Acme", primary_color="#1F4E79")
    assert b.primary_color == "#1F4E79"


def test_deliverable_defaults_to_linchpin_branding():
    d = Deliverable(title="T", client="C", summary="s")
    assert d.branding == DEFAULT_BRANDING
    assert "Prepared by Kern" in d.to_markdown()


def test_deliverable_markdown_shows_custom_branding_footer_not_linchpin():
    d = Deliverable(title="T", client="C", summary="s",
                    branding=Branding(name="Acme Consulting"))
    md = d.to_markdown()
    assert "Prepared by Acme Consulting" in md
    assert "Prepared by Kern" not in md


def test_deliverable_markdown_renders_a_logo_image_tag_when_configured():
    d = Deliverable(title="T", client="C", summary="s",
                    branding=Branding(name="Acme", logo_url="https://acme.example/logo.png"))
    md = d.to_markdown()
    assert "![Acme](https://acme.example/logo.png)" in md
    # the logo line must come before the title, i.e. it's a real header
    assert md.index("![Acme]") < md.index("# T - C")


def test_deliverable_markdown_omits_logo_tag_when_not_configured():
    md = Deliverable(title="T", client="C", summary="s").to_markdown()
    assert "![" not in md


def test_deliverable_excel_summary_sheet_shows_branding(tmp_path):
    d = Deliverable(title="T", client="C", summary="s",
                    branding=Branding(name="Acme", logo_url="https://acme.example/logo.png"))
    wb = load_workbook(d.to_excel(tmp_path / "d.xlsx"))
    rows = [tuple(r) for r in wb["Summary"].iter_rows(values_only=True)]
    # A distinct label from "Prepared" (the date row) so the two rows -
    # a date vs. a company name - aren't confusable when skimming column A.
    assert ("Brand", "Acme") in rows
    assert ("Logo", "https://acme.example/logo.png") in rows


def test_deliverable_excel_summary_sheet_omits_logo_row_when_not_configured(tmp_path):
    d = Deliverable(title="T", client="C", summary="s")
    wb = load_workbook(d.to_excel(tmp_path / "d.xlsx"))
    rows = [tuple(r) for r in wb["Summary"].iter_rows(values_only=True)]
    assert not any(r[0] == "Logo" for r in rows if r)


def test_from_result_maps_jobresult_fields():
    @dataclass
    class _Guided:
        summary: str

    @dataclass
    class _Result:
        summary: str
        confidence: float
        citations: tuple
        guided: object

    res = _Result(
        summary="Did the thing.",
        confidence=0.8,
        citations=("Safety Stock - vandeput.pdf Ch.5",),
        guided=_Guided(summary="You must approve the PO."),
    )
    d = from_result(res, title="Report", client="Acme", prepared="2026-06-23",
                    kpis=(Kpi("Fill rate", "95%"),))
    assert d.summary == "Did the thing."
    assert d.confidence == 0.8
    assert d.citations == ("Safety Stock - vandeput.pdf Ch.5",)
    assert d.residual == "You must approve the PO."  # pulled from guided when not given
    assert d.kpis[0].name == "Fill rate"
    assert "## Coverage & handoff" in d.to_markdown()
