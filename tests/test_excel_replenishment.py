"""Tests for jobs/excel_replenishment_job.py — replenish a client's planilla.

Mirrors the odoo_replenishment shape: prepare reads the system of record (here
the client's Excel file), run plans the restock and STAGES the write-back as a
dry-run changeset through the safe-staging plane, and the outcome is >=2 ranked
executable options. Nothing is ever written without an approval + apply.
"""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from jobs import excel_replenishment_job as job
from scm_agent import llm, tools
from scm_agent.orchestrator import Orchestrator
from src import writeback
from src.guided import ESCALATED, OPTIONS, passed_guided

SHEET = "Stock Bodega"


def _make_planilla(path, *, demand_column=False, order_column=False):
    """Client-style planilla: title row, Spanish headers at row 3, own formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "INVENTARIO BODEGA CENTRAL"
    headers = ["Codigo", "Descripcion", "Stock", "Punto Reorden"]
    if demand_column:
        headers.append("Demanda Semanal")
    if order_column:
        headers.append("Pedir (Kern)")
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    rows = [
        ("SKU-001", "Tornillo 3mm", 42, 50, 10.0),
        ("SKU-002", "Tuerca 3mm", 130, 80, 12.0),
        ("SKU-003", "Arandela", 8, 25, 4.0),
    ]
    for r, (code, desc, stock, rop, demand) in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=stock)
        ws.cell(row=r, column=4, value=rop)
        if demand_column:
            ws.cell(row=r, column=5, value=demand)
    wb.save(path)
    return path


@pytest.fixture
def planilla(tmp_path):
    return _make_planilla(tmp_path / "planilla.xlsx")


def _make_planilla_with_cost(path, unit_costs):
    """Client-style planilla WITH an optional unit-cost column - for
    financial-threshold escalation tests. Same SKU/stock/ROP as ``_make_planilla``."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "INVENTARIO BODEGA CENTRAL"
    headers = ["Codigo", "Descripcion", "Stock", "Punto Reorden", "Costo Unitario"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    rows = [("SKU-001", "Tornillo 3mm", 42, 50), ("SKU-002", "Tuerca 3mm", 130, 80), ("SKU-003", "Arandela", 8, 25)]
    for r, (code, desc, stock, rop) in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=stock)
        ws.cell(row=r, column=4, value=rop)
        ws.cell(row=r, column=5, value=unit_costs[code])
    wb.save(path)
    return path


# ---- prepare: sheet + column auto-detection ---------------------------------------

def test_prepare_autodetects_sheet_and_spanish_columns(planilla):
    payload = job.prepare(str(planilla), {})
    assert payload["sheet"] == SHEET
    assert payload["mode"] == "reorder-point"
    skus = [row.sku for row in payload["rows"]]
    assert skus == ["SKU-001", "SKU-002", "SKU-003"]
    assert payload["rows"][0].on_hand == 42
    assert payload["rows"][0].reorder_point == 50


def test_prepare_prefers_demand_mode_when_demand_column_present(tmp_path):
    p = _make_planilla(tmp_path / "d.xlsx", demand_column=True)
    payload = job.prepare(str(p), {})
    assert payload["mode"] == "demand-cover"
    assert payload["rows"][0].demand_per_period == 10.0


def test_prepare_respects_explicit_column_params(planilla):
    payload = job.prepare(str(planilla), {"sku_column": "Codigo", "stock_column": "Stock",
                                          "rop_column": "Punto Reorden", "sheet": SHEET})
    assert payload["sheet"] == SHEET
    assert len(payload["rows"]) == 3


def test_prepare_fails_clearly_without_sku_column(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Cosa", "Cantidad"])
    ws.append(["x", 1])
    f = tmp_path / "bad.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="SKU"):
        job.prepare(str(f), {})


def test_prepare_detects_an_optional_cost_column(tmp_path):
    p = _make_planilla_with_cost(tmp_path / "cost.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    payload = job.prepare(str(p), {})
    by_sku = {row.sku: row for row in payload["rows"]}
    assert by_sku["SKU-001"].cost == 5.0
    assert by_sku["SKU-003"].cost == 2.0


def test_prepare_cost_is_none_when_no_cost_column_present(planilla):
    payload = job.prepare(str(planilla), {})
    assert all(row.cost is None for row in payload["rows"])


def test_prepare_fails_clearly_without_rop_or_demand(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Codigo", "Stock"])
    ws.append(["SKU-1", 5])
    f = tmp_path / "no_signal.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="reorder point|demand"):
        job.prepare(str(f), {})


# ---- run: plan + staged changeset ---------------------------------------------------

def test_run_reorder_point_mode_orders_up_to_factor(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    # SKU-001: 42 < 50 -> order up to 2*50 => +58; SKU-002: 130 >= 80 -> 0;
    # SKU-003: 8 < 25 -> +42.
    assert report.restock == {"SKU-001": 58.0, "SKU-003": 42.0}
    assert report.n_restock == 2 and report.n_skus == 3
    assert report.total_restock == 100.0


def test_run_demand_cover_mode_targets_cover_periods(tmp_path):
    p = _make_planilla(tmp_path / "d.xlsx", demand_column=True)
    payload = job.prepare(str(p), {})
    report = job.run(payload, cover_periods=8.0)
    # SKU-001: target 80 vs 42 -> +38; SKU-002: 96 vs 130 -> 0; SKU-003: 32 vs 8 -> +24.
    assert report.restock == {"SKU-001": 38.0, "SKU-003": 24.0}
    assert report.mode == "demand-cover"


def test_run_stages_changeset_with_new_order_column(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    cs = report.changeset
    assert cs is not None and cs.risk_tier == writeback.TIER_REVERSIBLE
    edits = {c.field: c.after for c in cs.changes}
    # New column E: header at the header row + one qty per restocked SKU.
    assert edits["E3"] == "Pedir (Kern)"
    assert edits["E4"] == 58.0 and edits["E6"] == 42.0
    # Order-column cells are fresh writes (before None); the plan's INPUT cells
    # travel as no-op GUARDS (before == after) so input drift is caught at apply.
    by_field = {c.field: c for c in cs.changes}
    assert by_field["E4"].before is None
    assert by_field["C4"].before == 42 and by_field["C4"].after == 42   # stock guard
    assert by_field["D4"].before == 50 and by_field["D4"].after == 50   # ROP guard
    assert load_workbook(planilla)[SHEET]["E4"].value is None  # file untouched


def test_run_reuses_existing_order_column(tmp_path):
    p = _make_planilla(tmp_path / "o.xlsx", order_column=True)
    payload = job.prepare(str(p), {})
    report = job.run(payload)
    edits = {c.field: c.after for c in report.changeset.changes}
    assert "E3" not in edits  # header already exists -> not re-written
    assert edits["E4"] == 58.0


# -- financial-threshold escalation: a big-$ restock needs finance sign-off ---


def test_no_cost_column_never_escalates_regardless_of_threshold(planilla):
    """The base fixture has no cost column - the $ value can't be known, so this
    must degrade to plain options (never guess), not silently gate on 0."""
    payload = job.prepare(str(planilla), {})

    report = job.run(payload, financial_threshold=0.01)

    assert report.outcome.status == OPTIONS


def test_default_threshold_leaves_a_modest_priced_plan_as_plain_options(tmp_path):
    p = _make_planilla_with_cost(tmp_path / "c.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    payload = job.prepare(str(p), {})

    report = job.run(payload)  # restock value: 58*5 + 42*2 = 374 - well under the 50k default

    assert report.outcome.status == OPTIONS


def test_restock_value_above_threshold_escalates_to_finance(tmp_path):
    p = _make_planilla_with_cost(tmp_path / "c.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    payload = job.prepare(str(p), {})

    report = job.run(payload, financial_threshold=100.0)  # 374 > 100

    assert report.outcome.status == ESCALATED
    assert report.outcome.escalation.route_to
    assert report.outcome.escalation.sla
    assert len(report.outcome.escalation.options) >= 2      # the ranked options are NOT lost
    assert job.verify(report) == []


def test_escalated_deck_states_the_requirement_in_words(tmp_path):
    """The data model being correct (outcome.status == ESCALATED) is not the same
    guarantee as a human reading the ACTUAL rendered document ever seeing it."""
    p = _make_planilla_with_cost(tmp_path / "c.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    payload = job.prepare(str(p), {})
    report = job.run(payload, financial_threshold=100.0)
    assert report.outcome.status == ESCALATED  # sanity

    md = job.build_deck(report, client="Acme", confidence=0.85).to_markdown()

    assert "ESCALATED" in md
    assert "finance" in md.lower()
    assert report.outcome.escalation.sla in md


def test_escalated_apply_howto_leads_with_a_stop_warning(tmp_path):
    """apply_howto.md is the exact document an operator opens right before
    running the one-shot apply command - it must not silently hand over that
    recipe with zero mention that finance sign-off is required first."""
    p = _make_planilla_with_cost(tmp_path / "c.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    payload = job.prepare(str(p), {})
    report = job.run(payload, financial_threshold=100.0)

    written = job.write_operational(report, tmp_path / "out", "Acme")

    text = written["apply_howto"].read_text(encoding="utf-8")
    assert text.startswith("# STOP")
    assert "finance" in text.lower()
    assert "apply_replenishment.py" in text  # the recipe is still there, just gated


def test_non_escalated_apply_howto_has_no_stop_warning(planilla):
    """The routine (non-escalated) case must NOT regress: no spurious warning
    on an ordinary small plan."""
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)

    written = job.write_operational(report, planilla.parent / "out", "Acme")

    text = written["apply_howto"].read_text(encoding="utf-8")
    assert not text.startswith("# STOP")


def test_escalated_run_still_reaches_the_orchestrator_deck_with_visible_options(tmp_path):
    """End-to-end through Orchestrator.run(): the ranked options must reach the
    written deck's 'Options to act' section even when the outcome is escalated -
    not just live in report.outcome.escalation.options, unread by anything."""
    p = _make_planilla_with_cost(tmp_path / "c.xlsx", {"SKU-001": 5.0, "SKU-002": 3.0, "SKU-003": 2.0})
    orch = Orchestrator(registry=tools.build_default_registry(), provider=llm.RulesFallback())

    res = orch.run(
        "excel replenishment: update my excel planilla",
        data_path=str(p), client="Acme", out_dir=tmp_path / "out",
        overrides={"financial_threshold": 100.0},
    )

    assert res.status == "ok" and res.tool == "excel_replenishment"
    assert res.guided is not None and res.guided.status == ESCALATED
    assert len(res.guided.options) >= 2          # visible at the top level, not just inside escalation
    deck_path = Path(res.deliverables["deck_report"])
    md = deck_path.read_text(encoding="utf-8")
    assert "ESCALATED" in md
    assert "## Options to act" in md             # the options actually rendered, not just returned


def test_standalone_orchestrator_run_never_uses_optimized_targets(planilla, tmp_path):
    """A tool run through the orchestrator (not a package) has no derive_params
    mechanism upstream of it - optimized_targets can only ever be None here, so
    the run must be byte-identical to a plain job.run() call. Proves the
    no-regression guarantee at the real caller seam, not just the job's own
    default argument. JobResult exposes no raw report object, so this reads
    the operational CSV deliverable - the same file a real caller would read."""
    orch = Orchestrator(registry=tools.build_default_registry(), provider=llm.RulesFallback())
    res = orch.run("excel replenishment: update my excel planilla",
                   data_path=str(planilla), client="Acme", out_dir=tmp_path / "out")
    assert res.status == "ok" and res.tool == "excel_replenishment"
    csv_text = Path(res.deliverables["csv"]).read_text(encoding="utf-8")
    assert "kern_optimized" not in csv_text
    assert csv_text.count("client_sheet") == 3  # all 3 demo SKUs
    assert "58.0" in csv_text and "42.0" in csv_text  # same numbers as always


def test_all_above_target_never_escalates_even_with_a_tiny_threshold(tmp_path):
    """Every SKU deeply above its reorder point -> nothing to restock -> no dollar
    value at risk -> must never gate the 'hold' outcome behind finance."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    headers = ["Codigo", "Stock", "Punto Reorden", "Costo Unitario"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.append(["SKU-001", 9999, 10, 500.0])
    p = tmp_path / "deep.xlsx"
    wb.save(p)
    payload = job.prepare(str(p), {})

    report = job.run(payload, financial_threshold=0.01)

    assert report.restock == {}
    assert report.outcome.status == OPTIONS
    assert report.outcome.options[0].label.startswith("Hold")


def test_run_no_restock_needed_stages_nothing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-1", 100, 10])
    f = tmp_path / "full.xlsx"
    wb.save(f)
    report = job.run(job.prepare(str(f), {}))
    assert report.restock == {} and report.changeset is None
    assert report.outcome.status == OPTIONS  # still a protected, ranked outcome


# ---- outcome contract ----------------------------------------------------------------

def test_outcome_offers_ranked_executable_options(planilla):
    report = job.run(job.prepare(str(planilla), {}))
    out = report.outcome
    assert out.status == OPTIONS
    assert len(out.options) >= 2
    assert sum(1 for o in out.options if o.recommended) == 1
    assert all(o.action for o in out.options)
    assert passed_guided(out)


# ---- the loop actually closes: approve + apply the staged changeset -------------------

def test_staged_changeset_applies_to_the_real_file_with_approval(planilla):
    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    store = payload["store"]
    approval = writeback.approve(report.changeset, "operator")
    result = writeback.apply(store, report.changeset, approval=approval)
    assert result.applied
    ws = load_workbook(planilla)[SHEET]
    assert ws["E3"].value == "Pedir (Kern)"
    assert ws["E4"].value == 58.0
    assert ws["A1"].value == "INVENTARIO BODEGA CENTRAL"  # client content intact
    # And it is rollback-able, honoring the writeback contract end to end.
    store.rollback(report.changeset.idempotency_key)
    assert load_workbook(planilla)[SHEET]["E4"].value is None


# ---- verify / deliverables -------------------------------------------------------------

def test_verify_passes_on_good_report(planilla):
    report = job.run(job.prepare(str(planilla), {}))
    assert job.verify(report) == []


def test_verify_flags_empty_planilla(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    f = tmp_path / "empty.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="no usable SKU rows"):
        job.prepare(str(f), {})


# ---- adversarial-review regressions ------------------------------------------------

def test_blank_rop_row_is_excluded_not_zeroed(tmp_path):
    # A blank ROP must never coalesce to 0: with negative stock that would place a
    # spurious order, and with positive stock it silently never replenishes.
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-OK", 8, 25])
    ws.append(["SKU-BLANK", -5, None])
    f = tmp_path / "blank_rop.xlsx"
    wb.save(f)
    report = job.run(job.prepare(str(f), {}))
    assert report.restock == {"SKU-OK": 42.0}      # no spurious order for SKU-BLANK
    assert report.n_unplanned == 1
    assert "NOT planned" in report.summary          # surfaced, never silent


def test_duplicate_skus_fail_closed(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", 8, 25])
    ws.append(["SKU-001", 3, 25])
    f = tmp_path / "dup.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="duplicate SKU"):
        job.prepare(str(f), {})


def test_all_formula_stock_gives_actionable_error(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", "=Z1+Z2", 25])  # formula with no cached value
    f = tmp_path / "formulas.xlsx"
    wb.save(f)
    with pytest.raises(ValueError, match="formula"):
        job.prepare(str(f), {})


def test_corrupt_xlsx_raises_value_error_not_library_internals(tmp_path):
    f = tmp_path / "fake.xlsx"
    f.write_bytes(b"this is not a zip archive at all")
    with pytest.raises(ValueError, match="could not open"):
        job.prepare(str(f), {})


def test_column_binding_is_priority_ordered_not_hash_ordered(tmp_path):
    # Two stock-candidate labels in the same header row: "stock" outranks
    # "existencias" in the priority tuple, deterministically.
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Existencias", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", 999, 8, 25])
    f = tmp_path / "two_labels.xlsx"
    wb.save(f)
    payload = job.prepare(str(f), {})
    assert payload["rows"][0].on_hand == 8  # bound to "Stock", not "Existencias"


def test_sheet_scan_skips_leading_catalog_sheet(tmp_path):
    # A first sheet with a SKU column but no stock must not block a later sheet
    # that fully qualifies.
    wb = Workbook()
    catalog = wb.active
    catalog.title = "Catalogo"
    catalog.append(["Codigo", "Precio"])
    catalog.append(["SKU-001", 9.99])
    inv = wb.create_sheet(SHEET)
    inv.append(["Codigo", "Stock", "Punto Reorden"])
    inv.append(["SKU-001", 8, 25])
    f = tmp_path / "multi.xlsx"
    wb.save(f)
    payload = job.prepare(str(f), {})
    assert payload["sheet"] == SHEET


def test_idempotency_key_is_content_derived_and_stable(planilla):
    r1 = job.run(job.prepare(str(planilla), {}))
    r2 = job.run(job.prepare(str(planilla), {}))
    assert r1.changeset.idempotency_key == r2.changeset.idempotency_key  # same plan
    r3 = job.run(job.prepare(str(planilla), {}), order_up_to_factor=3.0)
    assert r3.changeset.idempotency_key != r1.changeset.idempotency_key  # new plan


def test_second_week_apply_does_not_collide_with_first(planilla):
    # Week 1: plan + apply. Week 2: stock moved, a NEW plan must apply cleanly
    # (no idempotent skip, no crash-window tripwire false positive).
    p1 = job.prepare(str(planilla), {})
    r1 = job.run(p1)
    writeback.apply(p1["store"], r1.changeset, approval=writeback.approve(r1.changeset, "op"))
    wb = load_workbook(planilla)
    wb[SHEET]["C4"] = 20  # week passes; stock drops further
    wb.save(planilla)
    p2 = job.prepare(str(planilla), {})
    r2 = job.run(p2)
    assert r2.changeset.idempotency_key != r1.changeset.idempotency_key
    result = writeback.apply(p2["store"], r2.changeset, approval=writeback.approve(r2.changeset, "op"))
    assert result.applied and not result.idempotent_skip
    assert load_workbook(planilla)[SHEET]["E4"].value == 80.0  # 2*50 - 20


def test_input_drift_between_stage_and_apply_is_refused(planilla):
    from src.connectors.excel import ExcelWritebackError

    payload = job.prepare(str(planilla), {})
    report = job.run(payload)
    wb = load_workbook(planilla)
    wb[SHEET]["C4"] = 49  # stock changed AFTER staging -> the plan's qty is stale
    wb.save(planilla)
    with pytest.raises(ExcelWritebackError, match="changed since staging"):
        writeback.apply(payload["store"], report.changeset,
                        approval=writeback.approve(report.changeset, "op"))
    assert load_workbook(planilla)[SHEET]["E4"].value is None  # nothing written


def test_apply_howto_deliverable_written_when_staged(planilla, tmp_path):
    report = job.run(job.prepare(str(planilla), {}))
    written = job.write_operational(report, tmp_path / "out", "Acme")
    howto = written["apply_howto"].read_text(encoding="utf-8")
    assert report.changeset.idempotency_key in howto
    assert "writeback.approve" in howto


def test_write_operational_emits_csv(planilla, tmp_path):
    report = job.run(job.prepare(str(planilla), {}))
    written = job.write_operational(report, tmp_path / "out", "Acme")
    assert written["csv"].exists()
    text = written["csv"].read_text(encoding="utf-8")
    assert "SKU-001" in text and "58" in text


def test_build_deck_writes_deliverable(planilla, tmp_path):
    report = job.run(job.prepare(str(planilla), {}))
    deck = job.build_deck(report, client="Acme", citations=("Vandeput (2020), ch. 2",))
    files = deck.write_all(tmp_path / "deck")
    assert any(p.exists() for p in files.values())


# ---- optimized-target override: prefer Kern's fresh (R,S) over the sheet's own ROP ---
#
# Wired from scm_agent/package_specs.py's _optimized_targets_from_inventory (a
# prior inventory_optimization step's report, within the same package run) --
# see tests/test_packages.py for the package-level wiring/degrade tests. These
# are job-level: standalone `run()` calls, so the mechanism is provable without
# spinning up a whole package.

def test_match_key_normalizes_openpyxl_float_skus():
    assert job._match_key("1001.0") == "1001"
    assert job._match_key(1001.0) == "1001"
    assert job._match_key("SKU-001") == "sku-001"
    assert job._match_key(" 42 ") == "42"


def test_match_key_folds_case():
    # ventas.csv and planilla.xlsx are independently client-authored -> a
    # casing mismatch between the two is plausible, not hypothetical.
    assert job._match_key("SKU-001") == job._match_key("sku-001")


def test_optimized_target_overrides_client_rop(planilla):
    payload = job.prepare(str(planilla), {})
    # Client's own sheet says SKU-002 (stock=130) is above its stale ROP=80 ->
    # old behavior would NOT restock. Kern's fresh optimized policy says the
    # real reorder point is higher -> it SHOULD restock, using Kern's numbers.
    optimized = {"SKU-002": {"reorder_point": 150.0, "order_up_to": 200.0}}
    report = job.run(payload, optimized_targets=optimized)
    line = next(ln for ln in report.lines if ln.sku == "SKU-002")
    assert line.source == "kern_optimized"
    assert line.trigger == 150.0
    assert line.target == 200.0
    assert line.restock_qty == 70.0  # 200 - 130, Kern's target, not the client's
    assert report.n_optimized == 1
    assert report.restock["SKU-002"] == 70.0


def test_optimized_derives_target_from_factor_when_order_up_to_none(planilla):
    # (s,Q) policies leave order_up_to unresolved at the job boundary (the R+Q
    # derivation itself lives in the package helper, tested separately) - the
    # job must still degrade gracefully to the factor-based target, never crash
    # or silently produce a None target.
    payload = job.prepare(str(planilla), {})
    optimized = {"SKU-001": {"reorder_point": 55.0, "order_up_to": None}}
    report = job.run(payload, optimized_targets=optimized, order_up_to_factor=3.0)
    line = next(ln for ln in report.lines if ln.sku == "SKU-001")
    assert line.source == "kern_optimized"
    assert line.target == 165.0  # 55 * 3.0


def test_sku_absent_from_optimizer_falls_back_to_client_sheet(planilla):
    payload = job.prepare(str(planilla), {})
    optimized = {"SKU-001": {"reorder_point": 60.0, "order_up_to": 100.0}}
    report = job.run(payload, optimized_targets=optimized)
    line = next(ln for ln in report.lines if ln.sku == "SKU-003")
    assert line.source == "client_sheet"
    assert line.trigger == 25.0    # the client's own ROP
    assert line.target == 50.0     # 25 * order_up_to_factor (default 2.0)
    assert report.n_optimized == 1  # only SKU-001


def test_prefer_optimized_false_ignores_targets(planilla):
    optimized = {"SKU-001": {"reorder_point": 999.0, "order_up_to": 999.0}}
    report = job.run(job.prepare(str(planilla), {}),
                     optimized_targets=optimized, prefer_optimized_policy=False)
    assert all(ln.source == "client_sheet" for ln in report.lines)
    assert report.n_optimized == 0
    baseline = job.run(job.prepare(str(planilla), {}))
    assert report.restock == baseline.restock


def test_run_without_optimized_targets_is_unchanged(planilla):
    """No-regression guarantee: the numbers must reproduce today's behavior
    exactly (see test_run_reorder_point_mode_orders_up_to_factor) when the
    caller passes nothing new."""
    report = job.run(job.prepare(str(planilla), {}))
    assert report.restock == {"SKU-001": 58.0, "SKU-003": 42.0}
    assert all(ln.source == "client_sheet" for ln in report.lines)
    assert all(ln.trigger == pytest.approx(row_rop) for ln, row_rop in
               zip(report.lines, (50.0, 80.0, 25.0)))
    assert report.n_optimized == 0


def test_demand_cover_sku_gets_trigger_when_optimized(tmp_path):
    p = _make_planilla(tmp_path / "d.xlsx", demand_column=True)
    payload = job.prepare(str(p), {})
    # SKU-002: stock=130, demand=12/wk. Blind demand-cover (8 periods) targets
    # 96 < 130 -> qty 0 regardless (today's behavior has no reorder TRIGGER at
    # all in demand-cover mode). An optimized R above current stock makes
    # Kern's min/max correctly fire a reorder that blind top-up would miss.
    optimized = {"SKU-002": {"reorder_point": 150.0, "order_up_to": 200.0}}
    report = job.run(payload, optimized_targets=optimized)
    line = next(ln for ln in report.lines if ln.sku == "SKU-002")
    assert line.source == "kern_optimized"
    assert line.trigger == 150.0
    assert line.restock_qty == 70.0  # 200 - 130, triggered because 130 < R=150
    other = next(ln for ln in report.lines if ln.sku == "SKU-001")
    assert other.source == "client_sheet"
    assert other.trigger is None  # demand-cover mode has no trigger concept today


def test_numeric_sku_matches_optimizer_key_across_formats(tmp_path):
    # openpyxl round-trips a whole-number numeric cell back as a clean "1001"
    # (verified empirically - no ".0"). The realistic mismatch runs the OTHER
    # direction: pandas upcasts an integer product_id COLUMN to float64 the
    # moment any other row in that CSV column is blank/NaN, so the optimizer's
    # key can arrive as "1001.0" even though the planilla's own SKU is clean.
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append([1001, 10, 5])
    f = tmp_path / "numeric_sku.xlsx"
    wb.save(f)
    payload = job.prepare(str(f), {})
    assert payload["rows"][0].sku == "1001"  # the planilla side is clean
    optimized = {"1001.0": {"reorder_point": 20.0, "order_up_to": 40.0}}  # pandas float64 key
    report = job.run(payload, optimized_targets=optimized)
    assert report.lines[0].source == "kern_optimized"


def test_write_operational_has_target_source_and_reorder_point_used(planilla, tmp_path):
    optimized = {"SKU-001": {"reorder_point": 60.0, "order_up_to": 100.0}}
    report = job.run(job.prepare(str(planilla), {}), optimized_targets=optimized)
    written = job.write_operational(report, tmp_path / "out", "Acme")
    text = written["csv"].read_text(encoding="utf-8")
    assert "target_source" in text and "reorder_point_used" in text
    assert "kern_optimized" in text


def test_deck_reports_partial_optimized_coverage(planilla):
    # Only SKU-001 has an optimized policy; SKU-002/003 fall back - partial
    # coverage must be named in the deck, never silently blended in.
    optimized = {"SKU-001": {"reorder_point": 60.0, "order_up_to": 100.0}}
    report = job.run(job.prepare(str(planilla), {}), optimized_targets=optimized)
    assert report.n_optimized == 1 and report.n_skus == 3
    md = job.build_deck(report, client="Acme").to_markdown()
    assert "Optimized coverage" in md
    assert "1/3" in md


def test_deck_omits_coverage_finding_when_fully_optimized(planilla):
    optimized = {
        "SKU-001": {"reorder_point": 60.0, "order_up_to": 100.0},
        "SKU-002": {"reorder_point": 150.0, "order_up_to": 200.0},
        "SKU-003": {"reorder_point": 30.0, "order_up_to": 60.0},
    }
    report = job.run(job.prepare(str(planilla), {}), optimized_targets=optimized)
    assert report.n_optimized == report.n_skus == 3
    md = job.build_deck(report, client="Acme").to_markdown()
    assert "used the sheet's own reorder point" not in md


def test_optimized_target_plans_even_when_client_column_is_blank(tmp_path):
    """Before this diff, a blank ROP/demand cell always hit n_unplanned += 1
    and could never reach restock/staging (the optimized-check now runs BEFORE
    that guard). An optimizer match must plan and stage the line anyway - the
    optimized target doesn't depend on the client's own column at all."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", 10, None])  # blank ROP -> today, always n_unplanned
    f = tmp_path / "blank_rop_optimized.xlsx"
    wb.save(f)
    payload = job.prepare(str(f), {})
    optimized = {"SKU-001": {"reorder_point": 50.0, "order_up_to": 80.0}}
    report = job.run(payload, optimized_targets=optimized)
    assert report.n_unplanned == 0
    line = report.lines[0]
    assert line.source == "kern_optimized"
    assert line.restock_qty == 70.0  # 80 - 10, staged
    assert report.restock == {"SKU-001": 70.0}
    assert report.changeset is not None
    edits = {c.field: c.after for c in report.changeset.changes}
    assert edits  # non-empty staged changeset - the blank ROP never blocked it


def test_verify_rejects_bad_source_label(planilla):
    report = job.run(job.prepare(str(planilla), {}))
    bad_line = replace(report.lines[0], source="mystery")
    bad_report = replace(report, lines=(bad_line,) + report.lines[1:])
    issues = job.verify(bad_report)
    assert any("source" in i for i in issues)
