"""Tests for src/connectors/excel.py — safe-staging writeback against an Excel file.

The store implements the same read/applied_keys/claim/release/commit/rollback
surface as writeback.InMemoryStore / the Odoo connector stores, so the entire
safety plane (tiers, signed approval, idempotency, audit) is reused unchanged.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from src import writeback
from src.connectors.excel import ExcelWorkbookStore, ExcelWritebackError
from src.writeback_store import SqliteAuditLedger

SHEET = "Stock Bodega"


def _make_planilla(path):
    """A client-style workbook: own title, headers at row 3 (not 1), own formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "INVENTARIO BODEGA CENTRAL"
    for col, h in enumerate(["Codigo", "Descripcion", "Stock", "Punto Reorden"], 1):
        ws.cell(row=3, column=col, value=h)
    rows = [("SKU-001", "Tornillo 3mm", 42, 50), ("SKU-002", "Tuerca 3mm", 130, 80)]
    for r, (code, desc, stock, rop) in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=stock)
        ws.cell(row=r, column=4, value=rop)
        ws.cell(row=r, column=5, value=f"=C{r}*100")  # client's own formula, must survive
    wb.save(path)
    return path


@pytest.fixture
def planilla(tmp_path):
    return _make_planilla(tmp_path / "planilla.xlsx")


def _stage(store, edits, key="stage-1", tier=writeback.TIER_REVERSIBLE):
    return writeback.stage(store, "excel:planilla.xlsx", {SHEET: edits},
                           risk_tier=tier, idempotency_key=key)


def _approved_apply(store, cs, who="operator"):
    approval = writeback.approve(cs, who)
    return writeback.apply(store, cs, approval=approval)


# ---- read ----------------------------------------------------------------------

def test_read_returns_cell_values_for_sheet(planilla):
    store = ExcelWorkbookStore(planilla)
    cells = store.read(SHEET)
    assert cells["A4"] == "SKU-001"
    assert cells["C4"] == 42
    assert cells["E4"] == "=C4*100"  # formulas read as content, not computed values


def test_read_missing_sheet_returns_empty(planilla):
    assert ExcelWorkbookStore(planilla).read("No Existe") == {}


def test_missing_file_raises_clear_error(tmp_path):
    store = ExcelWorkbookStore(tmp_path / "nope.xlsx")
    with pytest.raises(ExcelWritebackError, match="not found"):
        store.read(SHEET)


# ---- row -> cell resolution ------------------------------------------------------

def test_resolve_row_edits_finds_header_row_and_cells(planilla):
    store = ExcelWorkbookStore(planilla)
    cells = store.resolve_row_edits(SHEET, "Codigo", {"SKU-002": {"Punto Reorden": 90}})
    assert cells == {"D5": 90}


def test_resolve_row_edits_unknown_column_lists_available(planilla):
    store = ExcelWorkbookStore(planilla)
    with pytest.raises(ExcelWritebackError, match="Stock"):
        store.resolve_row_edits(SHEET, "Codigo", {"SKU-001": {"No Such Col": 1}})


def test_resolve_row_edits_unknown_row_key_raises(planilla):
    store = ExcelWorkbookStore(planilla)
    with pytest.raises(ExcelWritebackError, match="SKU-999"):
        store.resolve_row_edits(SHEET, "Codigo", {"SKU-999": {"Stock": 1}})


def test_resolve_row_edits_missing_key_column_raises(planilla):
    store = ExcelWorkbookStore(planilla)
    with pytest.raises(ExcelWritebackError, match="header"):
        store.resolve_row_edits(SHEET, "Columna Inexistente", {"SKU-001": {"Stock": 1}})


# ---- stage -> approve -> apply ---------------------------------------------------

def test_stage_builds_before_after_changeset(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90, "F4": "URGENTE"})
    by_field = {c.field: c for c in cs.changes}
    assert by_field["D5"].before == 80 and by_field["D5"].after == 90
    assert by_field["F4"].before is None and by_field["F4"].after == "URGENTE"


def test_apply_without_approval_is_refused_and_file_untouched(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90})
    with pytest.raises(writeback.WritebackRefused):
        writeback.apply(store, cs)
    assert load_workbook(planilla)[SHEET]["D5"].value == 80


def test_approved_apply_writes_file_and_creates_backup(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90, "F4": "URGENTE"})
    result = _approved_apply(store, cs)
    assert result.applied and not result.idempotent_skip

    ws = load_workbook(planilla)[SHEET]
    assert ws["D5"].value == 90 and ws["F4"].value == "URGENTE"

    backups = list(planilla.parent.glob("*.linchpin-backup*"))
    assert len(backups) == 1
    assert load_workbook(backups[0])[SHEET]["D5"].value == 80  # backup holds the ORIGINAL


def test_second_apply_same_key_is_idempotent_skip(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90})
    _approved_apply(store, cs)
    again = writeback.apply(store, cs, approval=writeback.approve(cs, "operator"))
    assert again.idempotent_skip and not again.applied


def test_client_formulas_and_title_survive_commit(planilla):
    store = ExcelWorkbookStore(planilla)
    _approved_apply(store, _stage(store, {"D5": 90}))
    ws = load_workbook(planilla)[SHEET]
    assert ws["A1"].value == "INVENTARIO BODEGA CENTRAL"
    assert ws["E4"].value == "=C4*100"
    assert ws["E5"].value == "=C5*100"


# ---- rollback ---------------------------------------------------------------------

def test_rollback_restores_values_and_clears_added_cells(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90, "F4": "URGENTE"})
    _approved_apply(store, cs)
    store.rollback(cs.idempotency_key)
    ws = load_workbook(planilla)[SHEET]
    assert ws["D5"].value == 80
    assert ws["F4"].value is None  # cell that did not exist before is cleared
    with pytest.raises(KeyError):
        store.rollback(cs.idempotency_key)  # forgotten after rollback


# ---- drift detection (file changed between stage and apply) -----------------------

def test_drift_between_stage_and_apply_fails_closed(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90})
    # A human edits the same cell after staging.
    wb = load_workbook(planilla)
    wb[SHEET]["D5"] = 85
    wb.save(planilla)
    with pytest.raises(ExcelWritebackError, match="changed since staging"):
        writeback.apply(store, cs, approval=writeback.approve(cs, "operator"))
    assert load_workbook(planilla)[SHEET]["D5"].value == 85  # nothing written
    # The claim was released -> a re-staged changeset with the same key can proceed.
    cs2 = _stage(store, {"D5": 90})
    assert _approved_apply(store, cs2).applied


# ---- persistence (idempotency survives a process restart via the ledger) ----------

def test_ledger_persists_idempotency_across_store_instances(planilla, tmp_path):
    ledger_path = tmp_path / "audit.sqlite3"
    cs = None
    store1 = ExcelWorkbookStore(planilla, ledger=SqliteAuditLedger(ledger_path))
    cs = _stage(store1, {"D5": 90}, key="restart-safe")
    _approved_apply(store1, cs)

    store2 = ExcelWorkbookStore(planilla, ledger=SqliteAuditLedger(ledger_path))
    again = writeback.apply(store2, cs, approval=writeback.approve(cs, "operator"))
    assert again.idempotent_skip
    # And the second instance can still roll the change back.
    store2.rollback("restart-safe")
    assert load_workbook(planilla)[SHEET]["D5"].value == 80


# ---- failure modes -----------------------------------------------------------------

def test_locked_file_surfaces_actionable_error(planilla, monkeypatch):
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90})
    import src.connectors.excel as excel_mod
    monkeypatch.setattr(excel_mod.os, "replace",
                        lambda *a, **k: (_ for _ in ()).throw(PermissionError("locked")))
    with pytest.raises(ExcelWritebackError, match="open in Excel"):
        writeback.apply(store, cs, approval=writeback.approve(cs, "operator"))
    assert load_workbook(planilla)[SHEET]["D5"].value == 80  # original untouched


def test_backup_dir_override(planilla, tmp_path):
    backups = tmp_path / "respaldo"
    store = ExcelWorkbookStore(planilla, backup_dir=backups)
    _approved_apply(store, _stage(store, {"D5": 90}))
    assert list(backups.glob("*.linchpin-backup*"))
    assert not list(planilla.parent.glob("*.linchpin-backup*"))


def test_commit_unknown_sheet_fails_closed(planilla):
    store = ExcelWorkbookStore(planilla)
    cs = writeback.stage(store, "excel:planilla.xlsx", {"Hoja Fantasma": {"A1": 1}},
                         risk_tier=writeback.TIER_REVERSIBLE, idempotency_key="ghost")
    with pytest.raises(ExcelWritebackError, match="Hoja Fantasma"):
        writeback.apply(store, cs, approval=writeback.approve(cs, "operator"))


def test_resolve_row_edits_duplicate_row_key_fails_closed(planilla):
    wb = load_workbook(planilla)
    wb[SHEET]["A6"] = "SKU-001"  # duplicate of row 4's key
    wb.save(planilla)
    store = ExcelWorkbookStore(planilla)
    with pytest.raises(ExcelWritebackError, match="duplicate row key"):
        store.resolve_row_edits(SHEET, "Codigo", {"SKU-001": {"Stock": 1}})


def test_colliding_sanitized_keys_get_distinct_backups(planilla):
    # "rop 1" and "rop-1" sanitize to the same safe key; the content-hash suffix
    # must keep their backups separate.
    store = ExcelWorkbookStore(planilla)
    _approved_apply(store, _stage(store, {"D5": 90}, key="rop 1"))
    _approved_apply(store, _stage(store, {"D4": 60}, key="rop-1"))
    assert len(list(planilla.parent.glob("*.linchpin-backup*"))) == 2


# ---- crash window between file write and audit record (stale-backup tripwire) ------

def test_crash_between_save_and_audit_is_caught_on_retry(planilla, monkeypatch):
    # If the process dies AFTER os.replace but BEFORE the audit record, the write
    # is live but unaudited. A naive retry re-stages from the mutated file, bakes
    # the new value in as `before`, and would record a poisoned restore (rollback
    # then "restores" to the new value). The orphaned backup left by the crashed
    # attempt must act as a tripwire: commit refuses until the operator reconciles.
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90}, key="crashy")
    monkeypatch.setattr(writeback.AuditBookkeeping, "record",
                        lambda self, entry: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(RuntimeError):
        writeback.apply(store, cs, approval=writeback.approve(cs, "operator"))
    monkeypatch.undo()

    assert load_workbook(planilla)[SHEET]["D5"].value == 90  # write landed, unaudited
    cs2 = _stage(store, {"D5": 95}, key="crashy")  # different content -> different hash
    with pytest.raises(ExcelWritebackError, match="earlier apply attempt"):
        writeback.apply(store, cs2, approval=writeback.approve(cs2, "operator"))
    assert load_workbook(planilla)[SHEET]["D5"].value == 90  # tripwire wrote nothing


def test_same_key_same_content_retry_proceeds_after_backup_only_crash(planilla):
    # A hard kill between the backup copy and the file write leaves a backup whose
    # content-hash suffix MATCHES the retried changeset (file unchanged, so the
    # re-stage produces the identical changeset) - that retry must proceed.
    import shutil

    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90}, key="pre-save-crash")
    orphan = planilla.parent / store._backup_name("pre-save-crash", cs.content_hash)
    shutil.copy2(planilla, orphan)
    assert _approved_apply(store, cs).applied


def test_rollback_removes_backups_enabling_clean_key_reuse(planilla):
    # After a successful rollback the backup is redundant (file restored) - keeping
    # it would false-trigger the tripwire on a legitimate reuse of the same key.
    store = ExcelWorkbookStore(planilla)
    cs = _stage(store, {"D5": 90}, key="reuse-me")
    _approved_apply(store, cs)
    store.rollback("reuse-me")
    assert not list(planilla.parent.glob("*.linchpin-backup*"))
    cs2 = _stage(store, {"D5": 91}, key="reuse-me")
    assert _approved_apply(store, cs2).applied
