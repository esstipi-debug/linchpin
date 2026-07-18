"""Commercial packages: spec integrity, the package-level QA gate, and end-to-end
demo runs for all eight packages (diagnostico / starter / growth / scale /
retainer_ejecutivo / proyecto_red_almacen / proyecto_sourcing / liquidacion)."""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import pytest

from examples.run_package import DEMO_PARAMS, build_demo_intake
from jobs.inventory_optimization import JobReport, SkuRecommendation
from scm_agent.package_specs import (
    DIAGNOSTICO,
    GROWTH,
    LIQUIDACION,
    PACKAGES,
    PROYECTO_RED_ALMACEN,
    PROYECTO_SOURCING,
    RETAINER_EJECUTIVO,
    SCALE,
    STARTER,
    _optimized_targets_from_inventory,
    get_package,
)
from scm_agent.packages import PackageSpec, PackageStep, missing_required_inputs, run_package
from scm_agent.registry import Prepared, Produced, Tool, ToolRegistry
from scm_agent.tools import build_default_registry
from src import client_profile
from src.deliverable import Branding


class _NoKnowledge:
    """Citation-free stand-in so tests never load the books graph."""

    def ground_citations(self, keywords, brief, limit=5):
        return []

    def ground_citations_detailed(self, keywords, brief, limit=5):
        return []

    def warnings(self):
        return []


@pytest.fixture(scope="module")
def demo_intake(tmp_path_factory) -> Path:
    return build_demo_intake(tmp_path_factory.mktemp("intake"))


def _run(spec, intake, out, **kwargs):
    kwargs.setdefault("params", dict(DEMO_PARAMS))
    kwargs.setdefault("knowledge", _NoKnowledge())
    kwargs.setdefault("clients_root", None)
    return run_package(spec, intake, out_dir=out, **kwargs)


# ---- spec integrity -----------------------------------------------------------

def test_every_step_tool_exists_in_registry():
    reg = build_default_registry()
    for spec in PACKAGES.values():
        for step in spec.steps:
            reg.get(step.tool_key)  # KeyError => spec drifted from the registry


def test_every_step_slot_exists_in_spec_inputs():
    for spec in PACKAGES.values():
        slots = {i.slot for i in spec.inputs}
        for step in spec.steps:
            if step.input_slot is not None:
                assert step.input_slot in slots, (spec.key, step.tool_key)


def test_no_duplicate_steps():
    for spec in PACKAGES.values():
        keys = spec.tool_keys()
        assert len(keys) == len(set(keys)), spec.key


def test_every_package_defaults_to_spanish():
    for spec in PACKAGES.values():
        assert spec.lang == "es", spec.key


def test_scope_matches_monetization_brief():
    """The brief (documentation/MONETIZATION_BRIEF.md) is the commercial source of
    truth: 4 tools in the diagnostic sprint, 8 in Starter, 26 in Growth (Starter +
    the diagnostic's E&O and financial KPIs + 16 more)."""
    assert set(DIAGNOSTICO.tool_keys()) == {
        "data_quality", "abc_xyz", "excess_obsolete", "financial_kpis",
    }
    assert set(STARTER.tool_keys()) == {
        "forecast", "abc_xyz", "whatif", "inventory_optimization", "newsvendor",
        "excel_replenishment", "cycle_count", "data_quality",
    }
    assert set(GROWTH.tool_keys()) == set(STARTER.tool_keys()) | {
        "excess_obsolete", "financial_kpis",
        "multi_echelon", "ddmrp", "simulation", "drp", "odoo_replenishment",
        "reconciliation", "fefo", "sourcing", "landed_cost", "acceptance_sampling",
        "pricing", "cost_to_serve", "learning_curve", "returns", "risk", "dea",
    }
    assert len(GROWTH.tool_keys()) == 26

    scale_extra = {
        "sop", "facility_location", "transportation", "warehouse_layout",
        "slotting", "queuing", "scheduling", "earned_value", "leadership_chain",
    }
    assert set(SCALE.tool_keys()) == set(GROWTH.tool_keys()) | scale_extra
    assert len(SCALE.tool_keys()) == 35, "the full catalog is 35 tools, not 34"

    assert set(PROYECTO_RED_ALMACEN.tool_keys()) == {
        "facility_location", "transportation", "warehouse_layout",
        "slotting", "queuing", "scheduling",
    }
    assert set(PROYECTO_SOURCING.tool_keys()) == {
        "sourcing", "landed_cost", "acceptance_sampling",
    }


def test_retainer_ejecutivo_is_scale_same_tools_different_governance():
    """The brief is explicit: Retainer Ejecutivo has the SAME 35 tools as Scale -
    what differs is cadence/SLA, not capability."""
    assert set(RETAINER_EJECUTIVO.tool_keys()) == set(SCALE.tool_keys())
    assert RETAINER_EJECUTIVO.key != SCALE.key
    assert RETAINER_EJECUTIVO.price != SCALE.price
    assert RETAINER_EJECUTIVO.cadence != SCALE.cadence


def test_derive_steps_follow_their_source():
    for spec in PACKAGES.values():
        seen: set[str] = set()
        for step in spec.steps:
            if step.tool_key == "cycle_count" and step.derive is not None:
                assert "abc_xyz" in seen, f"{spec.key}: derive before its source"
            seen.add(step.tool_key)


def test_excel_replenishment_follows_inventory_optimization():
    """_optimized_targets_from_inventory depends on inventory_optimization's
    report already being in `reports` when excel_replenishment's derive_params
    runs - a future reshuffle of _STARTER_STEPS must not silently break that."""
    for spec in PACKAGES.values():
        keys = spec.tool_keys()
        if "inventory_optimization" in keys and "excel_replenishment" in keys:
            assert keys.index("inventory_optimization") < keys.index("excel_replenishment"), spec.key


def test_get_package_unknown_key():
    with pytest.raises(KeyError):
        get_package("no-such-package")


# ---- derive_params mechanism: report-sourced params, the sibling of derive
# (report -> input file) and params_from_input (file -> params) --------------------

def _fake_tool(key: str, run_report) -> Tool:
    """A minimal registrable Tool whose run() echoes back the params it
    received (as the report), so a test can assert on what actually reached
    it after every PackageStep augmenter ran."""
    return Tool(
        key=key, title=key, description="test double", intent_keywords=(),
        requires_data=False,
        prepare=lambda request, provider: Prepared(status="ok", payload=None),
        run=lambda payload, params: Produced(report=run_report(params), summary="ok"),
        qa=lambda report: [],
        deliver=lambda report, out_dir, client: {},
    )


def _synthetic_registry(*tools: Tool) -> ToolRegistry:
    reg = ToolRegistry()
    for t in tools:
        reg.register(t)
    return reg


def _run_synthetic(spec: PackageSpec, registry: ToolRegistry, **kwargs):
    kwargs.setdefault("params", {})
    kwargs.setdefault("knowledge", _NoKnowledge())
    kwargs.setdefault("clients_root", None)
    return run_package(spec, None, registry=registry, **kwargs)


def test_derive_params_merges_report_derived_params():
    registry = _synthetic_registry(
        _fake_tool("fake_source", lambda params: {"value": 42}),
        _fake_tool("fake_sink", lambda params: dict(params)),
    )
    spec = PackageSpec(
        key="synthetic", title="Synthetic", price="n/a", cadence="n/a", audience="n/a",
        inputs=(),
        steps=(
            PackageStep("fake_source", None),
            PackageStep("fake_sink", None,
                       derive_params=lambda reports: {"injected": reports["fake_source"]["value"]}),
        ),
    )
    result = _run_synthetic(spec, registry)
    sink_outcome = next(s for s in result.steps if s.tool_key == "fake_sink")
    assert sink_outcome.status == "ok"
    assert sink_outcome.report["injected"] == 42


def test_derive_params_empty_dict_is_noop():
    registry = _synthetic_registry(_fake_tool("fake_sink", lambda params: dict(params)))
    spec = PackageSpec(
        key="synthetic", title="Synthetic", price="n/a", cadence="n/a", audience="n/a",
        inputs=(),
        steps=(PackageStep("fake_sink", None, derive_params=lambda reports: {}),),
    )
    result = _run_synthetic(spec, registry, params={"caller_key": "unchanged"})
    outcome = result.steps[0]
    assert outcome.report == {"caller_key": "unchanged"}


def test_derive_params_absent_source_report_degrades():
    """The source step never ran (not in reports) - a well-behaved derive_params
    uses .get() and returns {} rather than raising; the sink step still runs on
    its own params, matching every other optional-input degrade in this runner."""
    registry = _synthetic_registry(_fake_tool("fake_sink", lambda params: dict(params)))
    spec = PackageSpec(
        key="synthetic", title="Synthetic", price="n/a", cadence="n/a", audience="n/a",
        inputs=(),
        steps=(PackageStep("fake_sink", None,
                          derive_params=lambda reports: (
                              {"from_source": reports["never_ran"]} if "never_ran" in reports else {}
                          )),),
    )
    result = _run_synthetic(spec, registry)
    outcome = result.steps[0]
    assert outcome.status == "ok"
    assert "from_source" not in outcome.report


def test_derive_params_exception_marks_step_error():
    def _boom(reports):
        raise ValueError("boom")
    registry = _synthetic_registry(_fake_tool("fake_sink", lambda params: dict(params)))
    spec = PackageSpec(
        key="synthetic", title="Synthetic", price="n/a", cadence="n/a", audience="n/a",
        inputs=(),
        steps=(PackageStep("fake_sink", None, derive_params=_boom),),
    )
    result = _run_synthetic(spec, registry)
    assert result.status == "error"
    assert result.steps[-1].messages == ("boom",)


def test_derive_params_overwrites_matching_caller_param_key():
    """Same precedence as the two existing report/file param-augmenters
    (extra_input_params, params_from_input): the derived value wins, matching
    _run_step's established "later augmenter overwrites" pattern - there is no
    legitimate reason a caller pre-sets a key a derive_params is meant to own."""
    registry = _synthetic_registry(_fake_tool("fake_sink", lambda params: dict(params)))
    spec = PackageSpec(
        key="synthetic", title="Synthetic", price="n/a", cadence="n/a", audience="n/a",
        inputs=(),
        steps=(PackageStep("fake_sink", None, derive_params=lambda reports: {"key": "derived"}),),
    )
    result = _run_synthetic(spec, registry, params={"key": "caller"})
    assert result.steps[0].report["key"] == "derived"


# ---- _optimized_targets_from_inventory: the (s,Q) vs (R,S) derivation itself -----

def _fake_recommendation(product_id, reorder_point, order_quantity=None, order_up_to=None):
    return SkuRecommendation(
        product_id=product_id, method="test", intermittent=order_up_to is not None,
        forecast=0.0, error_std=0.0, bias=0.0, mae=0.0,
        policy_kind="(R, S)" if order_up_to is not None else "(s, Q)",
        order_quantity=order_quantity, order_up_to=order_up_to,
        reorder_point=reorder_point, safety_stock=0.0, z_factor=0.0,
        service_level=0.95, unit_cost=1.0, lead_periods=1.0,
        cycle_investment=0.0, ss_investment=0.0, investment=0.0, status="ok",
    )


def _fake_job_report(*recommendations):
    return JobReport(
        recommendations=list(recommendations), params={}, requested_investment=0.0,
        cycle_floor=0.0, final_investment=0.0, safety_stock_scale=1.0,
        feasible=True, budget=None, n_skus=len(recommendations), n_at_risk=0, n_intermittent=0,
    )


def test_optimized_targets_derives_s_from_r_plus_q_for_sq_policy():
    """(s,Q) continuous-review policies never populate order_up_to
    (src/policies.py::continuous_review_sq always returns
    order_up_to_level=None) - the helper must derive S = R + Q itself."""
    rec = _fake_recommendation("SKU-001", reorder_point=50.0, order_quantity=30.0, order_up_to=None)
    targets = _optimized_targets_from_inventory({"inventory_optimization": _fake_job_report(rec)})
    assert targets["optimized_targets"]["SKU-001"] == {"reorder_point": 50.0, "order_up_to": 80.0}


def test_optimized_targets_passes_through_rs_order_up_to():
    """(R,S) periodic-review policies DO populate order_up_to directly
    (src/policies.py::periodic_review_rs) - must be used as-is, never re-derived."""
    rec = _fake_recommendation("SKU-002", reorder_point=40.0, order_quantity=None, order_up_to=90.0)
    targets = _optimized_targets_from_inventory({"inventory_optimization": _fake_job_report(rec)})
    assert targets["optimized_targets"]["SKU-002"] == {"reorder_point": 40.0, "order_up_to": 90.0}


def test_optimized_targets_empty_when_optimizer_did_not_run():
    assert _optimized_targets_from_inventory({}) == {}


def test_optimized_targets_empty_when_optimizer_produced_no_recommendations():
    assert _optimized_targets_from_inventory(
        {"inventory_optimization": _fake_job_report()}
    ) == {}


# ---- intake gating ------------------------------------------------------------

def test_missing_required_inputs_lists_the_checklist(tmp_path):
    missing = missing_required_inputs(DIAGNOSTICO, tmp_path)
    assert len(missing) == 4
    assert any(line.startswith("maestro.csv") for line in missing)
    assert any("product_id, cogs, avg_inventory_value" in line for line in missing)


def test_empty_intake_blocks_package_and_writes_nothing(tmp_path):
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, tmp_path / "empty", out)
    assert result.status == "needs_data"
    assert result.deliverables == {}
    assert len(result.missing_inputs) == 4
    assert not (out / "diagnostico").exists()


def test_optional_steps_skip_without_blocking(demo_intake, tmp_path):
    """Growth without any optional file still delivers: optional steps skip."""
    partial = tmp_path / "partial"
    partial.mkdir()
    for name in ("ventas.csv", "maestro.csv", "planilla.xlsx", "supuestos.csv",
                 "stock.csv", "finanzas.csv", "pedidos.csv"):
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    params = {k: v for k, v in DEMO_PARAMS.items() if k != "use_odoo"}
    result = _run(GROWTH, partial, tmp_path / "out", params=params)
    assert result.status == "ok"
    skipped = {s.tool_key for s in result.steps if s.status == "skipped"}
    assert "reconciliation" in skipped and "odoo_replenishment" in skipped
    executed = {s.tool_key for s in result.steps if s.status == "ok"}
    assert {"forecast", "pricing", "cost_to_serve"} <= executed
    # the coverage table records why each optional step was skipped
    odoo = next(s for s in result.steps if s.tool_key == "odoo_replenishment")
    assert "Odoo" in odoo.skip_reason


# ---- the package-level QA gate --------------------------------------------------

def _registry_with_failing(key: str) -> ToolRegistry:
    reg = build_default_registry()
    broken = ToolRegistry()
    for tool in reg.list():
        if tool.key == key:
            tool = replace(tool, qa=lambda report: ["forced QA failure (test)"])
        broken.register(tool)
    return broken


def test_one_failing_qa_blocks_the_whole_package(demo_intake, tmp_path):
    """The per-tool guarantee, lifted to the package: ONE tool failing QA means
    NOTHING is written - not even the deliverables of the tools that passed."""
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, demo_intake, out,
                  registry=_registry_with_failing("financial_kpis"))
    assert result.status == "qa_failed"
    assert result.deliverables == {}
    assert any("forced QA failure" in issue for issue in result.qa_issues)
    assert not (out / "diagnostico").exists(), "QA failed but files were written"


def test_optional_step_qa_failure_also_blocks(demo_intake, tmp_path):
    """An optional step that RAN and failed QA blocks too: a package is one
    coherent deliverable. The escape hatch is removing that input and rerunning."""
    out = tmp_path / "out"
    result = _run(GROWTH, demo_intake, out, registry=_registry_with_failing("fefo"))
    assert result.status == "qa_failed"
    assert result.deliverables == {}
    assert not (out / "growth").exists()


def test_required_step_error_blocks_before_writing(demo_intake, tmp_path):
    reg = build_default_registry()
    broken = ToolRegistry()
    for tool in reg.list():
        if tool.key == "abc_xyz":
            def _boom(payload, params):
                raise RuntimeError("engine exploded (test)")
            tool = replace(tool, run=_boom)
        broken.register(tool)
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, demo_intake, out, registry=broken)
    assert result.status == "error"
    assert result.deliverables == {}
    assert not (out / "diagnostico").exists()


# ---- end-to-end on the demo intake ----------------------------------------------

def _assert_delivered(result, out: Path, spec_key: str, n_tools: int):
    assert result.status == "ok", (result.summary, result.qa_issues)
    executed = [s for s in result.steps if s.status == "ok"]
    assert len(executed) == n_tools
    root = out / spec_key
    assert (root / "deliverable.md").exists()
    assert (root / "deliverable.xlsx").exists()
    for step in executed:
        assert (root / step.tool_key).is_dir(), step.tool_key


def test_diagnostico_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, demo_intake, out)
    _assert_delivered(result, out, "diagnostico", 4)


# ---- E6: partner/white-label branding on the consolidated package deck -------

def test_package_deck_defaults_to_linchpin_branding(demo_intake, tmp_path):
    # DIAGNOSTICO defaults to lang="es" (see test_every_package_defaults_to_spanish)
    # so the footer label renders as "Preparado por", not "Prepared by".
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, demo_intake, out)
    assert result.status == "ok"
    deck = (out / "diagnostico" / "deliverable.md").read_text(encoding="utf-8")
    assert "Preparado por Kern" in deck


def test_package_deck_uses_the_clients_profile_branding_when_configured(demo_intake, tmp_path):
    root = tmp_path / "clients"
    client_profile.upsert_profile(
        "Acme Consulting", "Acme Consulting", root=root,
        branding=Branding(name="Acme Consulting", primary_color="#1F4E79"),
    )
    out = tmp_path / "out"
    result = _run(DIAGNOSTICO, demo_intake, out, client="Acme Consulting", clients_root=root)
    assert result.status == "ok"
    deck = (out / "diagnostico" / "deliverable.md").read_text(encoding="utf-8")
    assert "Preparado por Acme Consulting" in deck
    assert "Preparado por Kern" not in deck


def test_package_deck_explicit_branding_wins_over_the_clients_profile(demo_intake, tmp_path):
    root = tmp_path / "clients"
    client_profile.upsert_profile(
        "Acme Consulting", "Acme Consulting", root=root,
        branding=Branding(name="Acme Consulting"),
    )
    out = tmp_path / "out"
    result = _run(
        DIAGNOSTICO, demo_intake, out, client="Acme Consulting", clients_root=root,
        branding=Branding(name="Explicit Override Co"),
    )
    assert result.status == "ok"
    deck = (out / "diagnostico" / "deliverable.md").read_text(encoding="utf-8")
    assert "Preparado por Explicit Override Co" in deck
    assert "Prepared by Acme Consulting" not in deck


def test_diagnostico_deck_is_100pct_one_language_no_mixing(demo_intake, tmp_path):
    """E4 acceptance criterion: 'deck demo 100% en un solo idioma segun lang;
    cero mezcla' -- for the scope src/i18n.py actually covers (the consolidated
    deck's own labels + tool titles; see that module's docstring for what
    stays engine-native English regardless of lang by design)."""
    from src import i18n

    # Only fixed-text labels can be membership-checked directly (templated
    # ones like "{executed} of {total}..." render with numbers substituted).
    non_templated = [k for k, v in i18n.LABELS.items() if "{" not in v["es"] and "{" not in v["en"]]
    # Words short/common enough to collide with legitimate, out-of-scope
    # English tool-finding prose (e.g. the bare word "for") can't be used as
    # a reliable "other language leaked in" discriminator.
    too_common_to_discriminate = {
        "for_client", "cadence_word",
        "col_metric",  # "Metrica" (es) contains "Metric" (en) as a substring
    }
    # Excel-only labels (src.deliverable.Deliverable.to_excel) never render
    # into deliverable.md at all -- and some (hdr_client_field's "Client")
    # can coincidentally collide with unrelated deck content (here, the demo
    # run's client display name is literally "Client"). Covered separately
    # by test_diagnostico_deck_xlsx_headers_are_bilingual below.
    xlsx_only = {
        "hdr_title_field", "hdr_client_field", "sheet_summary", "sheet_kpis",
        "sheet_findings", "sheet_data_sources", "sheet_options", "sheet_citations",
        "col_finding", "col_detail", "col_impact", "col_option", "col_recommended",
        "col_summary", "yes_flag",
    }
    non_templated = [k for k in non_templated if k not in xlsx_only]

    decks = {}
    for lang in ("es", "en"):
        out = tmp_path / f"out_{lang}"
        spec = replace(DIAGNOSTICO, lang=lang)
        result = _run(spec, demo_intake, out)
        assert result.status == "ok"
        decks[lang] = (out / "diagnostico" / "deliverable.md").read_text(encoding="utf-8")

    for lang, other in (("es", "en"), ("en", "es")):
        deck = decks[lang]
        for key in non_templated:
            expected, unexpected = i18n.label(key, lang), i18n.label(key, other)
            if expected == unexpected:
                continue  # reads identically in both langs -- nothing to discriminate
            if expected not in deck and unexpected not in deck:
                continue  # this scenario never renders the label either way (e.g. nothing skipped)
            assert expected in deck, (lang, key, "expected label missing")
            if key not in too_common_to_discriminate:
                assert unexpected not in deck, (lang, key, "other-language label leaked in")

        for step in DIAGNOSTICO.steps:
            expected = i18n.tool_title(step.tool_key, lang)
            unexpected = i18n.tool_title(step.tool_key, other)
            assert expected in deck, (lang, step.tool_key)
            if expected != unexpected:
                assert unexpected not in deck, (lang, step.tool_key, "other-language title leaked in")


def test_diagnostico_deck_xlsx_headers_are_bilingual(demo_intake, tmp_path):
    """Complements test_diagnostico_deck_is_100pct_one_language_no_mixing for
    the Excel-only labels (sheet names, a few column headers) that never
    render into deliverable.md."""
    from openpyxl import load_workbook

    for lang, expect_sheet, unexpect_sheet in (("es", "Resumen", "Summary"), ("en", "Summary", "Resumen")):
        out = tmp_path / f"xlsx_{lang}"
        spec = replace(DIAGNOSTICO, lang=lang)
        result = _run(spec, demo_intake, out)
        assert result.status == "ok"
        wb = load_workbook(out / "diagnostico" / "deliverable.xlsx")
        assert expect_sheet in wb.sheetnames
        assert unexpect_sheet not in wb.sheetnames
        summary_sheet = wb[expect_sheet]
        first_col = [cell.value for cell in summary_sheet["A"] if cell.value is not None]
        expected_client_label = "Cliente" if lang == "es" else "Client"
        assert expected_client_label in first_col


def test_starter_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(STARTER, demo_intake, out)
    _assert_delivered(result, out, "starter", 8)
    # the consolidated deck names every executed tool in its coverage table --
    # translated (E4), since STARTER.lang defaults to "es"
    deck = (out / "starter" / "deliverable.md").read_text(encoding="utf-8")
    assert "Programa de Conteo Ciclico" in deck and "planilla.xlsx" in deck


def test_growth_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(GROWTH, demo_intake, out)
    _assert_delivered(result, out, "growth", 26)


def test_starter_threads_optimized_targets_into_replenishment(demo_intake, tmp_path):
    """The fix this section exists for: excel_replenishment must PREFER the
    prior inventory_optimization step's fresh (R,S) over the planilla's own
    stale reorder point/demand column, per SKU, within a package run."""
    out = tmp_path / "out"
    result = _run(STARTER, demo_intake, out)
    excel_step = next(s for s in result.steps if s.tool_key == "excel_replenishment")
    assert excel_step.status == "ok"
    # All 6 demo planilla SKUs have matching sales history in the demo ventas.csv.
    assert excel_step.report.n_optimized == 6
    assert all(ln.source == "kern_optimized" for ln in excel_step.report.lines)


def test_growth_and_scale_inherit_optimized_replenishment(demo_intake, tmp_path):
    """The wiring lives once on the shared _STARTER_STEPS tuple - GROWTH/SCALE
    inherit it automatically via _STARTER_STEPS + (...)."""
    for spec in (GROWTH, SCALE):
        out = tmp_path / f"out_{spec.key}"
        result = _run(spec, demo_intake, out)
        excel_step = next(s for s in result.steps if s.tool_key == "excel_replenishment")
        assert excel_step.status == "ok", spec.key
        assert all(ln.source == "kern_optimized" for ln in excel_step.report.lines), spec.key


def _demo_planilla_with_unmatched_sku(path):
    from openpyxl import Workbook

    from examples.run_package import _SKUS
    wb = Workbook()
    ws = wb.active
    ws.title = "Reposicion"
    ws.append(["SKU", "Stock Actual", "Demanda Semanal", "Pedido"])
    for i, sku in enumerate(_SKUS):
        ws.append([sku, 40 + 30 * i, 35 + 20 * i, ""])
    ws.append(["SKU-999", 10, 5, ""])  # no sales history -> optimizer has nothing for it
    wb.save(path)


def test_replenishment_degrades_when_optimizer_sku_missing(demo_intake, tmp_path):
    """Partial coverage - a planilla SKU with no sales history - must fall back
    to the client's own column for THAT line only, never block the package."""
    partial = tmp_path / "partial"
    partial.mkdir()
    for name in ("ventas.csv", "maestro.csv"):
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    _demo_planilla_with_unmatched_sku(partial / "planilla.xlsx")
    out = tmp_path / "out"
    result = _run(STARTER, partial, out)
    assert result.status == "ok"
    excel_step = next(s for s in result.steps if s.tool_key == "excel_replenishment")
    assert excel_step.status == "ok"
    lines_by_sku = {ln.sku: ln for ln in excel_step.report.lines}
    assert lines_by_sku["SKU-001"].source == "kern_optimized"
    assert lines_by_sku["SKU-999"].source == "client_sheet"
    assert excel_step.report.n_optimized == 6  # the 6 demo SKUs, not SKU-999


def test_whatif_fallback_template_runs_and_is_flagged(demo_intake, tmp_path):
    partial = tmp_path / "partial"
    partial.mkdir()
    for name in ("ventas.csv", "maestro.csv", "planilla.xlsx"):
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    out = tmp_path / "out"
    result = _run(STARTER, partial, out)
    assert result.status == "ok"
    whatif = next(s for s in result.steps if s.tool_key == "whatif")
    assert whatif.status == "ok"
    assert "rangos estandar" in whatif.summary
    deck = (out / "starter" / "deliverable.md").read_text(encoding="utf-8")
    assert "plantilla estandar" in deck


def test_cycle_count_derives_from_abc_classification(demo_intake, tmp_path):
    """The count program must count the same SKUs abc_xyz classified - the client
    never sends a second, pre-classified list."""
    out = tmp_path / "out"
    result = _run(STARTER, demo_intake, out)
    abc = next(s for s in result.steps if s.tool_key == "abc_xyz").report
    cc = next(s for s in result.steps if s.tool_key == "cycle_count").report
    assert cc.n_items == abc.n_skus
    abc_classes = {c.product_id: c.abc for c in abc.classifications}
    scheduled = {t.product_id for t in cc.schedule}
    assert scheduled <= set(abc_classes)


def test_scale_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(SCALE, demo_intake, out)
    _assert_delivered(result, out, "scale", 35)


def test_retainer_ejecutivo_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(RETAINER_EJECUTIVO, demo_intake, out)
    _assert_delivered(result, out, "retainer_ejecutivo", 35)


def test_proyecto_red_almacen_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(PROYECTO_RED_ALMACEN, demo_intake, out)
    _assert_delivered(result, out, "proyecto_red_almacen", 6)
    # warehouse_layout has no `deck=` (its output is layout.json/report.md/viewer,
    # not the standard Deliverable) - confirm it still delivers without one.
    files = {name for name in result.deliverables if name.startswith("warehouse_layout_")}
    assert files == {"warehouse_layout_layout", "warehouse_layout_report", "warehouse_layout_viewer"}


def test_proyecto_sourcing_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(PROYECTO_SOURCING, demo_intake, out)
    _assert_delivered(result, out, "proyecto_sourcing", 3)


def test_liquidacion_end_to_end(demo_intake, tmp_path):
    out = tmp_path / "out"
    result = _run(LIQUIDACION, demo_intake, out)
    # pricing is optional; the demo intake has ventas.csv (with price), so it runs too.
    _assert_delivered(result, out, "liquidacion", 4)
    liq = next(s for s in result.steps if s.tool_key == "markdown_liquidation")
    assert liq.report.total_recovered >= 0


def test_liquidacion_markdown_liquidation_uses_real_price_history_when_present(demo_intake, tmp_path):
    """Regression test: markdown_liquidation's price_history_path must come from
    the SAME ventas.csv the (separate) pricing step reads, wired via
    PackageStep.extra_input_params. Without that wiring the step silently falls
    back to default-markdown/salvage heuristics even when real price history is
    right there in the intake -- a >5x difference in recovered cash on this demo
    intake (caught by adversarial review before this test existed)."""
    out = tmp_path / "out"
    result = _run(LIQUIDACION, demo_intake, out)
    liq = next(s for s in result.steps if s.tool_key == "markdown_liquidation")
    assert liq.report.n_elasticity > 0
    assert liq.report.total_recovered > 40_000  # elasticity-priced, not the ~9.5k salvage-only figure


def test_liquidacion_markdown_liquidation_degrades_without_ventas_csv(demo_intake, tmp_path):
    partial = tmp_path / "partial"
    partial.mkdir()
    for name in ("maestro.csv", "stock.csv"):
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    out = tmp_path / "out"
    result = _run(LIQUIDACION, partial, out)
    liq = next(s for s in result.steps if s.tool_key == "markdown_liquidation")
    assert liq.report.n_elasticity == 0  # no ventas.csv -> the documented heuristic fallback


def test_liquidacion_pricing_step_is_optional(demo_intake, tmp_path):
    partial = tmp_path / "partial"
    partial.mkdir()
    for name in ("maestro.csv", "stock.csv"):
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    out = tmp_path / "out"
    result = _run(LIQUIDACION, partial, out)
    _assert_delivered(result, out, "liquidacion", 3)
    pricing_step = next(s for s in result.steps if s.tool_key == "pricing")
    assert pricing_step.status == "skipped"


def test_leadership_chain_scores_come_from_liderazgo_csv(demo_intake, tmp_path):
    """leadership_chain takes params["scores"], not a CSV data_path - confirm the
    package's params_from_input actually threads liderazgo.csv into the profile
    rather than silently falling through to needs_clarification."""
    out = tmp_path / "out"
    result = _run(SCALE, demo_intake, out)
    step = next(s for s in result.steps if s.tool_key == "leadership_chain")
    assert step.status == "ok"
    profile = step.report
    # liderazgo.csv fixture is C=3, H=2, A=3, I=1, N=2
    assert profile.scores == {"C": 3, "H": 2, "A": 3, "I": 1, "N": 2}


def test_warehouse_layout_uses_project_params_without_a_file(demo_intake, tmp_path):
    """warehouse_layout has input_slot=None (generative, not client-data-driven);
    confirm it runs off the step's static params, not the demo intake folder."""
    out = tmp_path / "out"
    result = _run(SCALE, demo_intake, out)
    step = next(s for s in result.steps if s.tool_key == "warehouse_layout")
    assert step.status == "ok"
    assert step.source == "conector (sin archivo)"
    layout, _report_md = step.report
    assert layout.building.width_m == 90.0
    assert len(layout.docks) == 10


def test_scale_skips_new_optional_tools_without_their_files(demo_intake, tmp_path):
    """Scale without any of the 9 new optional files still delivers on the Growth
    core - the new mando-ejecutivo tools skip, they don't block (same philosophy
    as test_optional_steps_skip_without_blocking for Growth)."""
    partial = tmp_path / "partial"
    partial.mkdir()
    core_files = (
        "ventas.csv", "maestro.csv", "planilla.xlsx", "supuestos.csv",
        "stock.csv", "finanzas.csv", "pedidos.csv",
    )
    for name in core_files:
        (partial / name).write_bytes((demo_intake / name).read_bytes())
    params = {k: v for k, v in DEMO_PARAMS.items() if k != "use_odoo"}
    result = _run(SCALE, partial, tmp_path / "out", params=params)
    assert result.status == "ok"
    skipped = {s.tool_key for s in result.steps if s.status == "skipped"}
    executed = {s.tool_key for s in result.steps if s.status == "ok"}
    # file-gated optional tools skip without their CSV
    file_gated_optional = {
        "facility_location", "transportation", "slotting",
        "queuing", "scheduling", "earned_value", "leadership_chain",
    }
    assert file_gated_optional <= skipped
    # warehouse_layout is optional too, but parametric (input_slot=None, no
    # file/gate) - it runs off its static step params regardless, so it's
    # never in "skipped"; sop is required (reuses ventas.csv, already present)
    assert {"sop", "warehouse_layout"} <= executed


def test_missing_required_inputs_for_proyecto_packages(tmp_path):
    missing_red = missing_required_inputs(PROYECTO_RED_ALMACEN, tmp_path)
    assert len(missing_red) == 5  # warehouse_layout has no file, so 6 tools -> 5 files
    assert any(line.startswith("ubicaciones.csv") for line in missing_red)
    assert any(line.startswith("trabajos.csv") for line in missing_red)

    missing_sourcing = missing_required_inputs(PROYECTO_SOURCING, tmp_path)
    assert len(missing_sourcing) == 3
    assert any(line.startswith("proveedores.csv") for line in missing_sourcing)
    assert any(line.startswith("calidad_aql.csv") for line in missing_sourcing)


def test_malformed_liderazgo_csv_blocks_with_an_actionable_message(demo_intake, tmp_path):
    """A client CSV with a score out of 0-4 or a missing CHAIN dimension column
    must not crash the runner, and must block the package (leadership_chain is
    optional, but an optional step that RAN and errored still blocks - same
    philosophy as test_optional_step_qa_failure_also_blocks) with a message an
    operator can act on, not a raw KeyError repr."""
    import pandas as pd

    partial = tmp_path / "partial"
    partial.mkdir()
    for name in demo_intake.iterdir():
        partial.joinpath(name.name).write_bytes(name.read_bytes())

    pd.DataFrame([{"C": 7, "H": 2, "A": 3, "I": 1, "N": 2}]).to_csv(
        partial / "liderazgo.csv", index=False
    )
    result = _run(SCALE, partial, tmp_path / "out_range")
    assert result.status == "error"
    assert result.deliverables == {}
    step = next(s for s in result.steps if s.tool_key == "leadership_chain")
    assert "fuera de 0-4" in step.messages[0]

    pd.DataFrame([{"C": 3, "H": 2, "A": 3, "I": 1}]).to_csv(
        partial / "liderazgo.csv", index=False
    )
    result = _run(SCALE, partial, tmp_path / "out_missing")
    assert result.status == "error"
    assert result.deliverables == {}
    step = next(s for s in result.steps if s.tool_key == "leadership_chain")
    assert "no tiene columna" in step.messages[0]


# ---- E4: lang threads into the LLM narrative rewrite --------------------------

class _RecordingProvider:
    """Available provider that records every prompt and returns a fixed reply."""

    def __init__(self):
        self.prompts: list[str] = []

    def available(self) -> bool:
        return True

    def complete(self, prompt: str) -> str:
        self.prompts.append(prompt)
        return "rewritten summary"

    def extract(self, prompt: str, schema: dict) -> dict:
        return {}


def test_package_step_narrative_uses_spec_lang(demo_intake, tmp_path):
    rec = _RecordingProvider()
    spec = replace(DIAGNOSTICO, lang="en")
    result = _run(spec, demo_intake, tmp_path / "out", provider=rec)
    assert result.status == "ok"
    assert any("English" in p for p in rec.prompts)
    assert not any("Spanish" in p for p in rec.prompts)
    executed = [s for s in result.steps if s.status == "ok"]
    assert all(s.summary == "rewritten summary" for s in executed)
