"""Tests for jobs/price_intelligence.py (Linchpin 3.0 PR-13 -- the sellable
one-shot milestone: intake refs -> acquire (PR-11 cascade) -> sanity (PR-12
gate) -> deliverable).

No real network call ever happens here -- every accepted/discarded row comes
from a frozen HTML fixture read via ``html_path`` (reusing PR-11's
``tests/fixtures/pricing_intel/*.html`` fixtures, plus one new fixture for
the unknown-currency discard path). The two domains used
(``example-retailer.test`` approved / ``example-blocked.test`` prohibited)
are PR-12's own committed synthetic ``config/sites/*.yaml`` fixtures.

Guarantees under test:
- prepare_records sniffs the refs columns and resolves html_path relative
  to the refs file's own directory;
- an end-to-end run over 3 refs produces 2 accepted + 1 discarded row, with
  the discard visible in the report (never silently dropped);
- the discarded/quarantined rows never leak into report.offers;
- jobs.qa.verify_price_intel/price_intel_passed enforce the >=60% coverage
  gate;
- write_deliverable produces the 3 named files (price_position_matrix.xlsx,
  report.md, ledger_export.csv) with real, E5-gated L3 citations in
  report.md;
- the deck respects a custom Branding (E6), exactly like every other tool's
  deliverable;
- the brief "donde estoy caro" resolves to the price_intelligence tool via
  intent classification.
"""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from jobs import price_intelligence as pi
from jobs.qa import price_intel_passed, verify_price_intel
from scm_agent import llm
from scm_agent.intent import classify
from scm_agent.tools import build_default_registry
from src.deliverable import Branding
from src.pricing_intel.ledger import PriceLedger

FIXTURES = Path(__file__).resolve().parent / "fixtures" / "pricing_intel"
FIXED_NOW = datetime(2026, 7, 12, 12, 0, 0, tzinfo=timezone.utc)


def _refs_df() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "product_id": "SKU-100", "competitor_url": "https://example-retailer.test/p/aw-3000",
            "our_price": 210.00, "html_path": "jsonld_clean.html",
        },
        {
            "product_id": "SKU-200", "competitor_url": "https://example-retailer.test/p/microdata-item",
            "our_price": 360.00, "html_path": "microdata_only.html",
        },
        {
            "product_id": "SKU-300", "competitor_url": "https://example-retailer.test/p/bad-currency",
            "our_price": 40.00, "html_path": "jsonld_unknown_currency.html",
        },
    ])


def _payload() -> dict:
    return pi.prepare_records(_refs_df(), base_dir=FIXTURES)


# -- prepare_records -----------------------------------------------------------


def test_prepare_records_sniffs_columns_and_resolves_html_path_relative_to_base_dir() -> None:
    payload = _payload()
    refs = payload["refs"]
    assert len(refs) == 3
    assert refs[0].product_id == "SKU-100"
    assert refs[0].site == "example-retailer.test"
    assert refs[0].our_price == Decimal("210.00") or refs[0].our_price == Decimal("210.0")
    assert Path(refs[0].html_path) == FIXTURES / "jsonld_clean.html"
    assert Path(refs[0].html_path).exists()


def test_prepare_records_derives_site_from_url_when_no_site_column() -> None:
    df = pd.DataFrame([{"product_id": "P1", "competitor_url": "https://shop.example.com/p/1"}])
    payload = pi.prepare_records(df)
    assert payload["refs"][0].site == "shop.example.com"


def test_prepare_records_id_only_ref_has_no_site() -> None:
    df = pd.DataFrame([{"product_id": "P1", "competitor_url": "MLA123456"}])
    payload = pi.prepare_records(df)
    assert payload["refs"][0].site is None


def test_prepare_raises_when_required_columns_missing() -> None:
    df = pd.DataFrame([{"foo": 1, "bar": 2}])
    with pytest.raises(ValueError, match="product_id"):
        pi.prepare_records(df)


# -- end-to-end run -------------------------------------------------------------


def test_run_end_to_end_accepts_two_and_discards_one_never_silently(tmp_path) -> None:
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(_payload(), ledger=ledger, event_ledger=None, now=FIXED_NOW)

    assert report.n_products == 3
    assert report.n_products_covered == 2
    assert report.coverage_pct == pytest.approx(2 / 3)
    assert len(report.offers) == 2
    assert {o.matched_product_id for o in report.offers} == {"SKU-100", "SKU-200"}

    # The bad-currency row is DISCARDED, not silently dropped -- it is
    # visible in report.rows/report.discarded with a machine-readable reason.
    assert len(report.discarded) == 1
    assert report.discarded[0].product_id == "SKU-300"
    assert report.discarded[0].reason == "unknown_currency"
    assert report.discarded[0].offer is None

    # Hand-verified extracted prices (see jsonld_clean.html / microdata_only.html
    # fixture docstrings, both USD so price_normalized == price identically).
    by_product = {o.matched_product_id: o for o in report.offers}
    assert by_product["SKU-100"].price == Decimal("199.99")
    assert by_product["SKU-100"].price_normalized == Decimal("199.99")
    assert by_product["SKU-100"].acquisition_tier == "L1"
    assert by_product["SKU-200"].price == Decimal("349.50")

    # Accepted offers are durably appended to the ledger (golden rule 8).
    record = ledger.latest_by_sku("example-retailer.test", "https://example-retailer.test/p/aw-3000")
    assert record is not None
    assert record.offer.price == Decimal("199.99")

    ledger.close()


def test_discarded_rows_never_leak_into_accepted_offers(tmp_path) -> None:
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(_payload(), ledger=ledger, event_ledger=None, now=FIXED_NOW)
    accepted_refs = {o.competitor_sku_ref for o in report.offers}
    assert "https://example-retailer.test/p/bad-currency" not in accepted_refs
    ledger.close()


# -- tool wiring: cross-thread sqlite3 hazard (scm_agent.tools._price_intelligence_run) ---
#
# POST /api/jobs dispatches tool.run() via asyncio.to_thread's default pool,
# which does not pin a given call to the same OS thread every time.
# jobs.price_intelligence.run() defaults its own `ledger` argument to
# src.pricing_intel.ledger.default_ledger() -- a process-wide CACHED sqlite3
# connection whose check_same_thread=True default binds it to whichever
# thread first touches it. Before the fix, scm_agent.tools._price_intelligence_run
# called run(payload) with no ledger at all, always falling through to that
# singleton -- a second call landing on a different pool thread than the one
# that first bound it raised sqlite3.ProgrammingError. The fix: own a fresh,
# call-scoped PriceLedger unless the caller supplied one via params.


def test_price_intelligence_run_never_touches_the_shared_default_singleton(tmp_path, monkeypatch) -> None:
    import functools

    import src.pricing_intel.ledger as ledger_module

    def _boom(*_a, **_kw):
        raise AssertionError("_price_intelligence_run must never touch the shared default singleton")

    monkeypatch.setattr(ledger_module, "default_ledger", _boom)
    # PriceLedger()'s own base_path default is def-time-bound (not re-read
    # per call), so isolate this test's fresh, owned instance into tmp_path
    # by patching the class name itself -- re-imported fresh on every
    # _price_intelligence_run call via its own local import statement.
    monkeypatch.setattr(
        ledger_module, "PriceLedger", functools.partial(ledger_module.PriceLedger, tmp_path / "owned_ledger"),
    )

    from scm_agent.tools import _price_intelligence_run

    produced = _price_intelligence_run(_payload(), {})

    assert produced.report.n_products == 3


def test_price_intelligence_run_survives_a_default_singleton_bound_to_a_different_thread(
    tmp_path, monkeypatch,
) -> None:
    """The real reproduction, mirroring test_price_watch_tool.py's equivalent
    test for _price_watch_run: bind default_ledger()'s singleton to a
    DIFFERENT OS thread first, then call _price_intelligence_run from THIS
    thread without a ledger override. Pre-fix this raised
    sqlite3.ProgrammingError every time (deterministic, not a race); post-fix
    it must succeed, because this call path no longer touches the singleton."""
    import threading

    import src.pricing_intel.ledger as ledger_module

    monkeypatch.setattr(ledger_module, "DEFAULT_BASE_PATH", tmp_path / "singleton_ledger")
    monkeypatch.setattr(ledger_module, "_default_ledger", None)
    # A plain functools.partial would collide with default_ledger()'s own
    # explicit-argument call (PriceLedger(DEFAULT_BASE_PATH)) below, so this
    # wrapper only substitutes tmp_path when called with NO argument -- see
    # test_price_watch_tool.py's equivalent test for the same note.
    _real_price_ledger = ledger_module.PriceLedger
    monkeypatch.setattr(
        ledger_module, "PriceLedger",
        lambda base_path=tmp_path / "owned_ledger": _real_price_ledger(base_path),
    )

    def _bind_singleton_on_another_thread():
        ledger_module.default_ledger()

    binder = threading.Thread(target=_bind_singleton_on_another_thread)
    binder.start()
    binder.join()

    from scm_agent.tools import _price_intelligence_run

    produced = _price_intelligence_run(_payload(), {})

    assert produced.report.n_products == 3


def test_site_not_approved_ref_is_skipped_and_visible_not_dropped(tmp_path) -> None:
    df = pd.concat([_refs_df(), pd.DataFrame([{
        "product_id": "SKU-400", "competitor_url": "https://example-blocked.test/p/whatever",
        "our_price": 10.0,
    }])], ignore_index=True)
    payload = pi.prepare_records(df, base_dir=FIXTURES)
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(payload, ledger=ledger, event_ledger=None, now=FIXED_NOW)

    assert report.n_products == 4
    assert report.n_products_covered == 2
    assert report.coverage_pct == pytest.approx(0.5)
    skipped_reasons = {r.product_id: r.reason for r in report.skipped}
    assert skipped_reasons["SKU-400"].startswith("site_not_approved")
    ledger.close()


# -- QA gate (jobs/qa.py) --------------------------------------------------------


def test_qa_passes_when_coverage_at_or_above_60_percent(tmp_path) -> None:
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(_payload(), ledger=ledger, event_ledger=None, now=FIXED_NOW)
    assert report.coverage_pct >= 0.60
    assert verify_price_intel(report) == []
    assert price_intel_passed(report) is True
    ledger.close()


def test_qa_blocks_when_coverage_below_60_percent(tmp_path) -> None:
    df = pd.concat([_refs_df(), pd.DataFrame([{
        "product_id": "SKU-400", "competitor_url": "https://example-blocked.test/p/whatever",
        "our_price": 10.0,
    }])], ignore_index=True)
    payload = pi.prepare_records(df, base_dir=FIXTURES)
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(payload, ledger=ledger, event_ledger=None, now=FIXED_NOW)

    assert report.coverage_pct == pytest.approx(0.5)
    issues = verify_price_intel(report)
    assert any("below the 60% minimum" in i for i in issues)
    assert price_intel_passed(report) is False
    ledger.close()


def test_qa_flags_a_freshness_sla_violation() -> None:
    report = pi.PriceIntelReport(
        n_products=1, n_products_covered=1, coverage_pct=1.0, offers=(), our_prices={},
        rows=(), quarantine_rate=0.0, avg_freshness_hours=100.0, sla_hours=48.0,
        tier_mix={}, stale_events=(), now=FIXED_NOW, summary="x",
    )
    issues = verify_price_intel(report)
    assert any("exceeds the 48.0h SLA" in i for i in issues)


# -- deliverable (report.md / price_position_matrix.xlsx / ledger_export.csv) --


def test_write_deliverable_produces_the_three_named_files_with_real_citations(tmp_path) -> None:
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(_payload(), ledger=ledger, event_ledger=None, now=FIXED_NOW)
    out_dir = tmp_path / "out"
    written = pi.write_deliverable(
        report, out_dir=out_dir, client="Acme",
        brief="donde estoy caro respecto a la competencia, analisis de posicion de precios",
        lang="es",
    )
    ledger.close()

    assert written["matrix"].exists()
    assert written["ledger_csv"].exists()
    assert written["report_md"].exists()
    assert written["report_md"].name == "report.md"
    assert written["matrix"].name == "price_position_matrix.xlsx"
    assert written["ledger_csv"].name == "ledger_export.csv"

    md = written["report_md"].read_text(encoding="utf-8")
    assert "Fuentes" in md  # golden rule 7's per-datum provenance section
    assert "SKU-100" in md
    assert "L1" in md
    # E5-gated L3 citations actually made it into the standard methodology section.
    assert "Metodologia" in md or "fundamento" in md
    assert "simon-price-management" in md  # a real, gated citation resolved

    wb = load_workbook(written["matrix"])
    assert "Position Matrix" in wb.sheetnames
    assert "Quarantine & Discards" in wb.sheetnames
    rows = [tuple(r) for r in wb["Quarantine & Discards"].iter_rows(values_only=True)]
    assert any(r[0] == "SKU-300" and r[3] == "discarded" for r in rows[1:])

    csv_text = written["ledger_csv"].read_text(encoding="utf-8")
    assert "SKU-100" in csv_text
    assert "SKU-300" not in csv_text  # discarded rows never enter the ledger export


def test_write_deliverable_respects_custom_branding_not_kern(tmp_path) -> None:
    ledger = PriceLedger(tmp_path / "ledger")
    report = pi.run(_payload(), ledger=ledger, event_ledger=None, now=FIXED_NOW)
    ledger.close()

    written = pi.write_deliverable(
        report, out_dir=tmp_path / "out", client="Acme", brief="price position",
        branding=Branding(name="Acme Consulting"), lang="en",
    )
    md = written["report_md"].read_text(encoding="utf-8")
    assert "Prepared by Acme Consulting" in md
    assert "Prepared by Kern" not in md


# -- intent classification -------------------------------------------------------


def test_donde_estoy_caro_resolves_to_price_intelligence() -> None:
    reg = build_default_registry()
    result = classify("donde estoy caro", reg, llm.RulesFallback())
    assert result.job_type == "price_intelligence"


def test_price_intelligence_registered_with_options_and_deck() -> None:
    reg = build_default_registry()
    tool = reg.get("price_intelligence")
    assert tool.requires_data is True
    assert tool.options is not None
    assert tool.deck is not None
