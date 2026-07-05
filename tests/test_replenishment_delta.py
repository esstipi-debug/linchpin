"""Tests for src/replenishment_delta.py + examples/monitor_planilla.py.

The recurring-service layer: snapshot a replenishment run, compare it with the
previous snapshot of the same planilla, and report WHAT CHANGED - new SKUs below
target, recovered ones, quantities that moved - instead of a full plan the client
has to re-read every week.
"""

from __future__ import annotations

import json

import pytest
from openpyxl import Workbook, load_workbook

from examples.monitor_planilla import main as monitor_main
from jobs import excel_replenishment_job as job
from src.replenishment_delta import compare, render_markdown, snapshot

SHEET = "Stock Bodega"


def _snap(lines):
    """A snapshot dict shaped like snapshot() output, from (sku, on_hand, qty) triples."""
    return {
        "version": 1,
        "label": "test",
        "filename": "p.xlsx",
        "sheet": SHEET,
        "mode": "reorder-point",
        "lines": [{"sku": s, "on_hand": oh, "target": 0.0, "restock_qty": q} for s, oh, q in lines],
    }


# ---- compare -------------------------------------------------------------------------

def test_compare_detects_new_orders_and_resolved():
    prev = _snap([("A", 100, 0.0), ("B", 5, 20.0)])
    curr = _snap([("A", 3, 40.0), ("B", 60, 0.0)])
    d = compare(prev, curr)
    assert d.new_orders == (("A", 40.0),)
    assert d.resolved == ("B",)
    assert d.still_short == 1
    assert d.has_changes


def test_compare_detects_quantity_moves():
    prev = _snap([("A", 10, 20.0), ("B", 10, 30.0)])
    curr = _snap([("A", 5, 35.0), ("B", 15, 10.0)])
    d = compare(prev, curr)
    assert d.qty_up == (("A", 20.0, 35.0),)
    assert d.qty_down == (("B", 30.0, 10.0),)
    assert d.new_orders == () and d.resolved == ()


def test_compare_detects_added_and_removed_skus():
    prev = _snap([("A", 10, 0.0)])
    curr = _snap([("A", 10, 0.0), ("NEW", 2, 8.0)])
    d = compare(prev, curr)
    assert d.added_skus == ("NEW",)
    assert d.new_orders == (("NEW", 8.0),)
    d2 = compare(curr, prev)
    assert d2.removed_skus == ("NEW",)


def test_compare_no_changes_is_quiet():
    prev = _snap([("A", 10, 20.0)])
    d = compare(prev, prev)
    assert not d.has_changes
    assert "sin cambios" in d.summary.lower() or "no changes" in d.summary.lower()


def test_compare_rejects_unknown_snapshot_version():
    prev = _snap([("A", 10, 0.0)])
    bad = dict(prev, version=99)
    with pytest.raises(ValueError, match="version"):
        compare(bad, prev)


# ---- snapshot / render -----------------------------------------------------------------

def _make_planilla(path, stock_001=42):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Codigo", "Stock", "Punto Reorden"])
    ws.append(["SKU-001", stock_001, 50])
    ws.append(["SKU-002", 130, 80])
    wb.save(path)
    return path


def test_snapshot_round_trips_through_json(tmp_path):
    p = _make_planilla(tmp_path / "p.xlsx")
    report = job.run(job.prepare(str(p), {}))
    snap = snapshot(report, label="week-1")
    again = json.loads(json.dumps(snap))
    assert again["label"] == "week-1"
    assert {ln["sku"] for ln in again["lines"]} == {"SKU-001", "SKU-002"}


def test_render_markdown_contains_sections():
    prev = _snap([("A", 100, 0.0), ("B", 5, 20.0)])
    curr = _snap([("A", 3, 40.0), ("B", 60, 0.0)])
    md = render_markdown(compare(prev, curr), client="Acme")
    assert "A" in md and "40" in md
    assert "Acme" in md


# ---- monitor CLI -----------------------------------------------------------------------

def test_monitor_first_run_saves_baseline(tmp_path, capsys):
    p = _make_planilla(tmp_path / "p.xlsx")
    state = tmp_path / "state.json"
    rc = monitor_main(["--file", str(p), "--state", str(state), "--label", "week-1"])
    assert rc == 0
    assert state.exists()
    assert "baseline" in capsys.readouterr().out.lower()


def test_monitor_second_run_reports_delta(tmp_path, capsys):
    p = _make_planilla(tmp_path / "p.xlsx", stock_001=100)  # week 1: healthy
    state = tmp_path / "state.json"
    report_md = tmp_path / "delta.md"
    monitor_main(["--file", str(p), "--state", str(state), "--label", "week-1"])

    wb = load_workbook(p)
    wb[SHEET]["B2"] = 10  # week 2: SKU-001 fell below its ROP of 50
    wb.save(p)
    rc = monitor_main(["--file", str(p), "--state", str(state), "--label", "week-2",
                       "--report-out", str(report_md)])
    assert rc == 0
    out = capsys.readouterr().out
    assert "SKU-001" in out
    md = report_md.read_text(encoding="utf-8")
    assert "SKU-001" in md and "week-1" in md and "week-2" in md
    # State advanced: a third run against the unchanged file reports no changes.
    monitor_main(["--file", str(p), "--state", str(state), "--label", "week-3"])
    assert "no changes" in capsys.readouterr().out.lower()
